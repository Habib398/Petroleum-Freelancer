from __future__ import annotations

import datetime
import json
import csv
import io
import os
import zipfile
import tempfile
import shutil
from pathlib import Path
from flask import jsonify, render_template, request, Response

from db import get_conn
from services.brand import get_brand

def register(app):
    """Admin-only APIs used by the admin inbox, executive dashboard and audit view."""
    ctx = app.extensions["ctx"]
    login_required = ctx.login_required
    role_required = ctx.role_required

    def _audit_where_from_args(args):
        """Build WHERE/params for audit log filters."""
        q = (args.get("q") or "").strip()
        entity = (args.get("entity") or "").strip()
        action = (args.get("action") or "").strip()
        module = (args.get("module") or "").strip()
        actor = (args.get("actor_user_id") or "").strip()
        date_from = (args.get("from") or "").strip()
        date_to = (args.get("to") or "").strip()

        where = ["1=1"]
        params = []

        if entity:
            where.append("entity = ?")
            params.append(entity)
        if action:
            where.append("action = ?")
            params.append(action)
        if actor:
            where.append("actor_user_id = ?")
            try:
                params.append(int(actor))
            except Exception:
                params.append(actor)
        # module is stored inside meta_json for document uploads/deletes and some exports
        if module:
            where.append("meta_json LIKE ?")
            params.append(f'%"module": "{module}"%')

        if date_from:
            where.append("date(created_at) >= date(?)")
            params.append(date_from)
        if date_to:
            where.append("date(created_at) <= date(?)")
            params.append(date_to)

        if q:
            like = f"%{q}%"
            where.append("(action LIKE ? OR entity LIKE ? OR entity_id LIKE ? OR meta_json LIKE ?)")
            params.extend([like, like, like, like])

        return " AND ".join(where), params

    def _fetch_audit(args, limit: int = 300):
        where_sql, params = _audit_where_from_args(args)
        conn = get_conn(); cur = conn.cursor()
        cur.execute(
            f"""
            SELECT id, actor_user_id, action, entity, entity_id, meta_json, created_at
            FROM audit_log
            WHERE {where_sql}
            ORDER BY id DESC
            LIMIT ?
            """
            , tuple(params + [int(limit)])
        )
        rows = cur.fetchall()
        conn.close()
        return rows


    @app.get("/api/search")
    @login_required
    def api_global_search():
        """Global search across key entities.
        - Admin: all stations
        - Others: limited to their station scope
        """
        me = ctx.get_me() or {}
        q = (request.args.get("q") or "").strip()
        if not q or len(q) < 2:
            return jsonify({"ok": True, "q": q, "results": []})

        like = f"%{q}%"
        brand = get_brand()
        conn = get_conn(); cur = conn.cursor()

        scope_station_ids = None
        if me.get("role") != "admin":
            scope_station_ids = list(ctx.station_scope_ids(me))
            if not scope_station_ids:
                scope_station_ids = [-9999]

        results = []

        # Stations (admin only)
        if me.get("role") == "admin":
            cur.execute("SELECT id, code, name FROM stations WHERE brand=? AND (code LIKE ? OR name LIKE ?) LIMIT 20", (brand, like, like))
            for r in cur.fetchall():
                results.append({"type": "station", "id": r["id"], "title": f"{r['code']} - {r['name']}", "url": f"/admin/stations"})

        # Users
        if me.get("role") == "admin":
            cur.execute("SELECT id, username, role FROM users WHERE username LIKE ? LIMIT 20", (like,))
        else:
            in_clause = ",".join(["?"] * len(scope_station_ids))
            cur.execute(f"SELECT id, username, role FROM users WHERE station_id IN ({in_clause}) AND username LIKE ? LIMIT 20", tuple(scope_station_ids + [like]))
        for r in cur.fetchall():
            results.append({"type": "user", "id": r["id"], "title": f"{r['username']} ({r['role']})", "url": f"/admin/users"})

        # Activities
        cur.execute("SELECT id, title FROM activities WHERE brand=? AND title LIKE ? LIMIT 20", (brand, like))
        for r in cur.fetchall():
            results.append({"type": "activity", "id": r["id"], "title": r["title"], "url": "/admin/activities"})

        # Documents (current)
        if scope_station_ids is None:
            cur.execute("SELECT id, module, section, title FROM documents WHERE brand=? AND is_current=1 AND (title LIKE ? OR section LIKE ?) ORDER BY id DESC LIMIT 20", (brand, like, like))
        else:
            in_clause = ",".join(["?"] * len(scope_station_ids))
            cur.execute(f"SELECT id, module, section, title FROM documents WHERE brand=? AND is_current=1 AND (station_id IS NULL OR station_id IN ({in_clause})) AND (title LIKE ? OR section LIKE ?) ORDER BY id DESC LIMIT 20",
                        tuple([brand] + scope_station_ids + [like, like]))
        for r in cur.fetchall():
            results.append({"type": "doc", "id": r["id"], "title": f"[{r['module']}] {r['title']}", "url": f"/admin/{(r['module'] or 'sasisopa')}"})

        conn.close()
        return jsonify({"ok": True, "q": q, "results": results[:60]})

    # (calibraciones UI moved to routes/calibraciones.py)





    @app.get("/api/admin/backup.zip")
    @login_required
    @role_required("admin")
    def api_admin_backup_zip():
        """Download a full backup (SQLite DB + uploads folder)."""
        me = ctx.get_me() or {}
        base_dir = Path(__file__).resolve().parent.parent
        db_path = base_dir / "data" / "cog_work_log.db"
        uploads_dir = base_dir / "uploads"

        tmpdir = tempfile.mkdtemp(prefix="cog_backup_")
        zip_path = os.path.join(tmpdir, "cog_backup.zip")
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
            if db_path.exists():
                z.write(str(db_path), arcname="data/cog_work_log.db")
            if uploads_dir.exists():
                for root, _dirs, files in os.walk(str(uploads_dir)):
                    for fn in files:
                        ap = os.path.join(root, fn)
                        rel = os.path.relpath(ap, str(base_dir))
                        z.write(ap, arcname=rel.replace("\\","/"))

        ctx.log_action(me, "download_backup", "system", "backup", {})
        return Response(
            open(zip_path, "rb").read(),
            mimetype="application/zip",
            headers={"Content-Disposition": "attachment; filename=cog_backup.zip"},
        )

    @app.post("/api/admin/restore")
    @login_required
    @role_required("admin")
    def api_admin_restore():
        """Restore from a backup ZIP (admin only). Expects multipart form: file=<zip>."""
        me = ctx.get_me() or {}
        f = request.files.get("file")
        if not f or not (f.filename or "").lower().endswith(".zip"):
            return jsonify({"ok": False, "error": "missing_zip"}), 400

        base_dir = Path(__file__).resolve().parent.parent
        data_dir = base_dir / "data"
        uploads_dir = base_dir / "uploads"
        db_path = data_dir / "cog_work_log.db"

        tmpdir = Path(tempfile.mkdtemp(prefix="cog_restore_"))
        zip_file = tmpdir / "restore.zip"
        f.save(zip_file)

        try:
            with zipfile.ZipFile(zip_file, "r") as z:
                z.extractall(tmpdir / "x")
        except Exception:
            shutil.rmtree(tmpdir, ignore_errors=True)
            return jsonify({"ok": False, "error": "invalid_zip"}), 400

        extracted = tmpdir / "x"
        new_db = extracted / "data" / "cog_work_log.db"
        if not new_db.exists():
            shutil.rmtree(tmpdir, ignore_errors=True)
            return jsonify({"ok": False, "error": "db_not_found_in_zip"}), 400

        data_dir.mkdir(parents=True, exist_ok=True)
        try:
            shutil.copy2(new_db, db_path)
        except Exception as e:
            shutil.rmtree(tmpdir, ignore_errors=True)
            return jsonify({"ok": False, "error": "restore_failed", "message": str(e)}), 500

        # Merge uploads (copy in, do not delete existing)
        try:
            extracted_uploads = extracted / "uploads"
            if extracted_uploads.exists():
                uploads_dir.mkdir(parents=True, exist_ok=True)
                for root, _dirs, files in os.walk(str(extracted_uploads)):
                    for fn in files:
                        ap = Path(root) / fn
                        rel = ap.relative_to(extracted_uploads)
                        dest = uploads_dir / rel
                        dest.parent.mkdir(parents=True, exist_ok=True)
                        shutil.copy2(ap, dest)
        except Exception:
            pass

        ctx.log_action(me, "restore_backup", "system", "restore", {"file": f.filename})
        return jsonify({"ok": True})


    def _row_to_dict(r):
        return dict(r) if r is not None else None

    @app.get("/api/admin/executive")
    @login_required
    @role_required("admin")
    def api_admin_executive():
        """Executive summary: compliance by user and risk by station."""
        today = datetime.date.today()
        month_start = today.replace(day=1)

        conn = get_conn()
        cur = conn.cursor()

        # Users (exclude admin)
        cur.execute(
            """
            SELECT u.id AS user_id, u.username, u.role, u.station_id,
                   s.code AS station_code, s.name AS station_name
            FROM users u
            LEFT JOIN stations s ON s.id = u.station_id
            WHERE u.is_active=1 AND u.role != 'admin'
            ORDER BY u.username
            """
        )
        users = cur.fetchall()

        # Compliance = approved submissions vs total submissions in the month (for that user)
        compliance = []
        for u in users:
            cur.execute(
                """
                SELECT
                  SUM(CASE WHEN status='approved' THEN 1 ELSE 0 END) AS approved,
                  COUNT(*) AS total
                FROM submissions
                WHERE user_id = ? AND date(created_at) >= date(?) AND date(created_at) <= date(?)
                """,
                (u["user_id"], month_start.isoformat(), today.isoformat()),
            )
            agg = cur.fetchone()
            approved = int(agg["approved"] or 0)
            total = int(agg["total"] or 0)
            pct = round((approved / total * 100.0), 1) if total else 0.0
            compliance.append(
                {
                    "user_id": u["user_id"],
                    "username": u["username"],
                    "role": u["role"],
                    "station_id": u["station_id"],
                    "station_code": u["station_code"],
                    "station_name": u["station_name"],
                    "approved": approved,
                    "total": total,
                    "pct": pct,
                }
            )

        # Risk per station (last 30 days): overdue=open red alerts + rejected submissions + open red alerts
        since = (today - datetime.timedelta(days=30)).isoformat()
        cur.execute(
            """
            SELECT s.id AS station_id, s.code AS station_code, s.name AS station_name,
                   SUM(CASE WHEN a.severity='red' AND a.status='open' THEN 1 ELSE 0 END) AS red_open
            FROM stations s
            LEFT JOIN alerts a ON a.station_id = s.id AND date(a.created_at) >= date(?)
            WHERE COALESCE(s.monthly_status,'active') IN ('active','view_only')
            GROUP BY s.id
            ORDER BY s.code
            """,
            (since,),
        )
        stations = cur.fetchall()

        risk = []
        for st in stations:
            cur.execute(
                """
                SELECT
                  SUM(CASE WHEN status='rejected' THEN 1 ELSE 0 END) AS rejected,
                  SUM(CASE WHEN status IN ('submitted','reviewed') THEN 1 ELSE 0 END) AS pending
                FROM submissions
                WHERE station_id = ? AND date(created_at) >= date(?)
                """,
                (st["station_id"], since),
            )
            sa = cur.fetchone()
            rejected = int(sa["rejected"] or 0)
            pending = int(sa["pending"] or 0)
            red_open = int(st["red_open"] or 0)

            score = red_open * 3 + rejected * 2 + pending * 1
            level = "low"
            if score >= 15:
                level = "high"
            elif score >= 6:
                level = "medium"

            risk.append(
                {
                    "station_id": st["station_id"],
                    "station_code": st["station_code"],
                    "station_name": st["station_name"],
                    "red_open": red_open,
                    "rejected": rejected,
                    "pending": pending,
                    "score": score,
                    "level": level,
                }
            )

        return jsonify({"today": today.isoformat(), "month_start": month_start.isoformat(), "compliance": compliance, "risk": risk})

    @app.get("/api/admin/audit")
    @login_required
    @role_required("admin")
    def api_admin_audit():
        rows = _fetch_audit(request.args, limit=300)
        audit = []
        for r in rows:
            d = dict(r)
            try:
                d["meta"] = json.loads(d.get("meta_json") or "{}")
            except Exception:
                d["meta"] = {}
            audit.append(d)
        return jsonify({"audit": audit})

    @app.get("/api/admin/audit/export.csv")
    @login_required
    @role_required("admin")
    def api_admin_audit_export_csv():
        rows = _fetch_audit(request.args, limit=5000)
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(["id","created_at","actor_user_id","action","entity","entity_id","meta_json"])
        for r in rows:
            w.writerow([r["id"], r["created_at"], r["actor_user_id"], r["action"], r["entity"], r["entity_id"], r["meta_json"]])
        payload = out.getvalue().encode("utf-8-sig")
        fname = "audit_export.csv"
        return Response(
            payload,
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'}
        )

    @app.get("/admin/sasisopa/historico/print")
    @login_required
    @role_required("admin")
    def admin_sasisopa_historico_print():
        rows = _fetch_audit(request.args, limit=1000)
        # parse meta for nicer display
        parsed = []
        for r in rows:
            d = dict(r)
            try:
                d["meta"] = json.loads(d.get("meta_json") or "{}")
            except Exception:
                d["meta"] = {}
            parsed.append(d)
        return render_template("sasisopa/historico_print.html", rows=parsed, args=request.args)

    @app.get("/api/admin/inbox")
    @login_required
    @role_required("admin")
    def api_admin_inbox():
        station_id = (request.args.get("station_id") or "").strip()
        date_from = (request.args.get("from") or "").strip()
        date_to = (request.args.get("to") or "").strip()
        severity = (request.args.get("severity") or "").strip()
        alert_status = (request.args.get("alert_status") or "").strip()
        submission_status = (request.args.get("submission_status") or "").strip()
        q = (request.args.get("q") or "").strip()

        limit = min(max(int(request.args.get("limit") or 250), 10), 1000)
        page = max(int(request.args.get("page") or 1), 1)
        offset = (page - 1) * limit

        conn = get_conn()
        cur = conn.cursor()

        brand = get_brand()

        # KPIs
        kpis = {}
        where_station = " AND station_id = ?" if station_id else ""
        params_station = [int(station_id)] if station_id else []

        cur.execute(
            f"SELECT COUNT(*) AS c FROM submissions WHERE brand=? AND status IN ('submitted','reviewed'){where_station}",
            [brand, *params_station],
        )
        kpis["submissions_pending"] = int(cur.fetchone()["c"] or 0)

        cur.execute(
            f"SELECT COUNT(*) AS c FROM payments WHERE brand=? AND status='pending'{where_station}",
            [brand, *params_station],
        )
        kpis["payments_pending"] = int(cur.fetchone()["c"] or 0)

        if station_id:
            cur.execute(
                "SELECT COUNT(*) AS c FROM alerts WHERE brand=? AND severity='red' AND status='open' AND station_id=?",
                (brand, int(station_id)),
            )
        else:
            cur.execute("SELECT COUNT(*) AS c FROM alerts WHERE brand=? AND severity='red' AND status='open'", (brand,))
        kpis["red_alerts"] = int(cur.fetchone()["c"] or 0)

        # Documental (SASISOPA / SGM): submissions pending review
        try:
            if station_id:
                cur.execute(
                    """
                    SELECT COUNT(*) AS c
                    FROM doc_submissions ds
                    JOIN doc_requirements dr ON dr.id=ds.requirement_id AND dr.brand=ds.brand AND dr.module=ds.module
                    WHERE ds.brand=? AND ds.module='sasisopa' AND ds.review_status='PENDING' AND dr.station_id=?
                    """,
                    (brand, int(station_id)),
                )
            else:
                cur.execute(
                    "SELECT COUNT(*) AS c FROM doc_submissions WHERE brand=? AND module='sasisopa' AND review_status='PENDING'",
                    (brand,),
                )
            kpis["sasisopa_pending"] = int(cur.fetchone()["c"] or 0)
        except Exception:
            kpis["sasisopa_pending"] = 0

        try:
            if station_id:
                cur.execute(
                    """
                    SELECT COUNT(*) AS c
                    FROM doc_submissions ds
                    JOIN doc_requirements dr ON dr.id=ds.requirement_id AND dr.brand=ds.brand AND dr.module=ds.module
                    WHERE ds.brand=? AND ds.module='sgm' AND ds.review_status='PENDING' AND dr.station_id=?
                    """,
                    (brand, int(station_id)),
                )
            else:
                cur.execute(
                    "SELECT COUNT(*) AS c FROM doc_submissions WHERE brand=? AND module='sgm' AND review_status='PENDING'",
                    (brand,),
                )
            kpis["sgm_pending"] = int(cur.fetchone()["c"] or 0)
        except Exception:
            kpis["sgm_pending"] = 0

        # Calibraciones: tanques incompletos (faltan documentos)
        try:
            if station_id:
                cur.execute(
                    """
                    SELECT COUNT(*) AS c
                    FROM cal_tanks
                    WHERE brand=? AND station_id=? AND (
                        pdf_path IS NULL OR TRIM(COALESCE(pdf_path,''))='' OR
                        sonda_pdf_path IS NULL OR TRIM(COALESCE(sonda_pdf_path,''))='' OR
                        temp_pdf_path IS NULL OR TRIM(COALESCE(temp_pdf_path,''))=''
                    )
                    """,
                    (brand, int(station_id)),
                )
            else:
                cur.execute(
                    """
                    SELECT COUNT(*) AS c
                    FROM cal_tanks
                    WHERE brand=? AND (
                        pdf_path IS NULL OR TRIM(COALESCE(pdf_path,''))='' OR
                        sonda_pdf_path IS NULL OR TRIM(COALESCE(sonda_pdf_path,''))='' OR
                        temp_pdf_path IS NULL OR TRIM(COALESCE(temp_pdf_path,''))=''
                    )
                    """,
                    (brand,),
                )
            kpis["calibraciones_incompletas"] = int(cur.fetchone()["c"] or 0)
        except Exception:
            kpis["calibraciones_incompletas"] = 0

        # Submissions list (latest 250)
        sub_where = ["1=1"]
        sub_params = []
        if station_id:
            sub_where.append("sub.station_id = ?"); sub_params.append(int(station_id))
        if submission_status:
            sub_where.append("sub.status = ?"); sub_params.append(submission_status)
        if date_from:
            sub_where.append("date(sub.created_at) >= date(?)"); sub_params.append(date_from)
        if date_to:
            sub_where.append("date(sub.created_at) <= date(?)"); sub_params.append(date_to)
        if q:
            like = f"%{q}%"
            sub_where.append("(a.title LIKE ? OR u.username LIKE ? OR s.name LIKE ? OR s.code LIKE ? OR sub.notes LIKE ?)")
            sub_params.extend([like, like, like, like, like])

        cur.execute(
            f"""
            SELECT sub.id, sub.event_id, sub.activity_id, sub.station_id, sub.user_id,
                   sub.notes, sub.evidence_path, sub.status, sub.created_at,
                   e.start_date AS event_date,
                   a.title AS activity_title,
                   u.username AS user_name,
                   s.code AS station_code, s.name AS station_name
            FROM submissions sub
            LEFT JOIN calendar_events e ON e.id = sub.event_id
            LEFT JOIN activities a ON a.id = sub.activity_id
            LEFT JOIN users u ON u.id = sub.user_id
            LEFT JOIN stations s ON s.id = sub.station_id
            WHERE sub.brand=? AND {' AND '.join(sub_where)}
            ORDER BY sub.id DESC
            LIMIT ? OFFSET ?
            """,
            [brand, *sub_params, limit, offset],
        )
        submissions = [dict(r) for r in cur.fetchall()]

        # Payments list (latest 250)
        pay_where = ["1=1"]
        pay_params = []
        if station_id:
            pay_where.append("p.station_id = ?"); pay_params.append(int(station_id))
        if date_from:
            pay_where.append("date(p.created_at) >= date(?)"); pay_params.append(date_from)
        if date_to:
            pay_where.append("date(p.created_at) <= date(?)"); pay_params.append(date_to)

        cur.execute(
            f"""
            SELECT p.id, p.station_id, p.period_start, p.period_end, p.status, p.proof_path, p.invoice_path, p.created_at,
                   s.code AS station_code, s.name AS station_name
            FROM payments p
            LEFT JOIN stations s ON s.id = p.station_id
            WHERE p.brand=? AND {' AND '.join(pay_where)}
            ORDER BY p.id DESC
            LIMIT ? OFFSET ?
            """,
            [brand, *pay_params, limit, offset],
        )
        payments = [dict(r) for r in cur.fetchall()]

        # Alerts list (latest 250)
        al_where = ["1=1"]
        al_params = []
        if station_id:
            al_where.append("a.station_id = ?"); al_params.append(int(station_id))
        if severity:
            al_where.append("a.severity = ?"); al_params.append(severity)
        if alert_status:
            al_where.append("a.status = ?"); al_params.append(alert_status)
        if date_from:
            al_where.append("date(a.created_at) >= date(?)"); al_params.append(date_from)
        if date_to:
            al_where.append("date(a.created_at) <= date(?)"); al_params.append(date_to)
        if q:
            like = f"%{q}%"
            al_where.append("(a.title LIKE ? OR a.description LIKE ? OR s.name LIKE ? OR s.code LIKE ?)")
            al_params.extend([like, like, like, like])

        cur.execute(
            f"""
            SELECT a.id, a.station_id, a.severity, a.title, a.description, a.status, a.created_at,
                   s.code AS station_code, s.name AS station_name
            FROM alerts a
            LEFT JOIN stations s ON s.id = a.station_id
            WHERE a.brand=? AND {' AND '.join(al_where)}
            ORDER BY a.id DESC
            LIMIT ? OFFSET ?
            """,
            [brand, *al_params, limit, offset],
        )
        alerts = [dict(r) for r in cur.fetchall()]

        # Activity overview (for the selected range): totals vs done/pending/rejected + missing list
        activity_overview = {"by_station": [], "missing": []}
        try:
            # Date range defaults if missing
            import datetime
            if not date_from:
                today = datetime.date.today()
                date_from = today.replace(day=1).isoformat()
            if not date_to:
                date_to = datetime.date.today().isoformat()

            # Stations scope
            if station_id:
                cur.execute("SELECT id, code, name FROM stations WHERE id=?", (int(station_id),))
            else:
                cur.execute("SELECT id, code, name FROM stations WHERE 1=1 ORDER BY code")
            st_rows = [dict(r) for r in cur.fetchall()]

            # For each station: total events in range, latest submission status per event
            for st in st_rows:
                sid = int(st["id"])

                # total events for station in range
                cur.execute(
                    "SELECT COUNT(*) AS c FROM calendar_events ce "
                    "WHERE ce.brand=? AND (ce.station_id IS NULL OR ce.station_id=?) "
                    "AND date(ce.start_date)>=date(?) AND date(ce.start_date)<=date(?)",
                    (brand, sid, date_from, date_to),
                )
                total = int(cur.fetchone()["c"] or 0)

                # latest submission per event (by max id)
                cur.execute(
                    """
                    SELECT sub.status, COUNT(*) AS c
                    FROM submissions sub
                    JOIN (
                      SELECT event_id, MAX(id) AS max_id
                      FROM submissions
                    WHERE brand=? AND station_id=?
                      GROUP BY event_id
                    ) last ON last.max_id=sub.id
                    JOIN calendar_events ce ON ce.id=sub.event_id
                    WHERE sub.station_id=?
                      AND (ce.station_id IS NULL OR ce.station_id=?)
                      AND ce.brand=?
                      AND date(ce.start_date)>=date(?) AND date(ce.start_date)<=date(?)
                    GROUP BY sub.status
                    """,
                    (brand, sid, sid, sid, brand, date_from, date_to),
                )
                by_status = {r["status"]: int(r["c"] or 0) for r in cur.fetchall()}
                rejected = by_status.get("rejected", 0)
                pending = by_status.get("submitted", 0) + by_status.get("reviewed", 0)
                approved = by_status.get("approved", 0)

                # done = approved + submitted (so operators see progress immediately)
                done = approved + by_status.get("submitted", 0)

                # Missing = total - any submissions (latest status counts)
                have = approved + pending + rejected
                missing = max(total - have, 0)

                activity_overview["by_station"].append(
                    {
                        "station_id": sid,
                        "station_code": st.get("code"),
                        "station_name": st.get("name"),
                        "total": total,
                        "done": done,
                        "pending": pending + missing,
                        "rejected": rejected,
                    }
                )

            # Build missing list (latest 250): events with no submission OR latest is rejected
            # Scope respects station filter.
            if station_id:
                st_filter = "=?"
                st_params = [int(station_id)]
            else:
                st_filter = "IN (SELECT id FROM stations)"
                st_params = []

            # We create a per-station row by expanding station list (manageable for inbox ranges).
            # Strategy: pull events for range, then for each station check latest submission.
            cur.execute(
                "SELECT id, start_date, title, repeat_kind, station_id "
                "FROM calendar_events WHERE brand=? AND date(start_date)>=date(?) AND date(start_date)<=date(?) "
                "ORDER BY date(start_date) DESC, id DESC LIMIT 800",
                (brand, date_from, date_to),
            )
            evs = [dict(r) for r in cur.fetchall()]

            # Station rows
            if station_id:
                cur.execute("SELECT id, code, name FROM stations WHERE id=?", (int(station_id),))
            else:
                cur.execute("SELECT id, code, name FROM stations ORDER BY code")
            stations_all = [dict(r) for r in cur.fetchall()]

            missing_items = []
            for ev in evs:
                for st in stations_all:
                    sid = int(st["id"])
                    # event applies to all or specific station
                    if ev.get("station_id") is not None and int(ev["station_id"]) != sid:
                        continue
                    cur.execute(
                        "SELECT status FROM submissions WHERE brand=? AND station_id=? AND event_id=? ORDER BY id DESC LIMIT 1",
                        (brand, sid, int(ev["id"])),
                    )
                    last = cur.fetchone()
                    if (not last) or (last["status"] in ("rejected",)):
                        missing_items.append(
                            {
                                "date": ev.get("start_date"),
                                "station_id": sid,
                                "station_code": st.get("code"),
                                "station_name": st.get("name"),
                                "activity_title": ev.get("title"),
                                "repeat_kind": ev.get("repeat_kind"),
                                "status": (last["status"] if last else "missing"),
                            }
                        )
                    if len(missing_items) >= 250:
                        break
                if len(missing_items) >= 250:
                    break
            activity_overview["missing"] = missing_items
        except Exception:
            # Don't break inbox if overview fails.
            activity_overview = {"by_station": [], "missing": []}

        return jsonify({"kpis": kpis, "submissions": submissions, "payments": payments, "alerts": alerts, "activity_overview": activity_overview, "page": page, "limit": limit})

    # ---------------- exports (Excel/PDF) ----------------

    @app.get("/api/admin/inbox/export.xlsx")
    @login_required
    @role_required("admin")
    def api_admin_inbox_export_xlsx():
        """Export current inbox (filtered) to an Excel file."""
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from flask import send_file

        # Reuse JSON payload from inbox
        payload = api_admin_inbox().get_json()
        wb = Workbook()

        def add_sheet(name: str, rows: list[dict]):
            ws = wb.create_sheet(title=name)
            if not rows:
                ws.append(["(sin datos)"])
                return
            headers = list(rows[0].keys())
            ws.append(headers)
            for r in rows:
                ws.append([r.get(h) for h in headers])
            # Auto width
            for i, h in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(i)].width = min(max(len(str(h)), 12), 40)

        # Remove default sheet
        wb.remove(wb.active)
        add_sheet("KPIs", [payload.get("kpis") or {}])
        add_sheet("Submissions", payload.get("submissions") or [])
        add_sheet("Payments", payload.get("payments") or [])
        add_sheet("Alerts", payload.get("alerts") or [])
        add_sheet("Overview_ByStation", payload.get("activity_overview", {}).get("by_station") or [])
        add_sheet("Overview_Missing", payload.get("activity_overview", {}).get("missing") or [])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="inbox_export.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.get("/api/admin/inbox/export.pdf")
    @login_required
    @role_required("admin")
    def api_admin_inbox_export_pdf():
        """Export a lightweight PDF summary of the inbox."""
        from io import BytesIO
        from flask import send_file
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas

        payload = api_admin_inbox().get_json()
        buf = BytesIO()
        c = canvas.Canvas(buf, pagesize=letter)
        width, height = letter
        y = height - 50

        def line(txt: str, dy: int = 14):
            nonlocal y
            c.drawString(40, y, txt[:120])
            y -= dy
            if y < 60:
                c.showPage()
                y = height - 50

        line("COG / Inbox Export")
        line(" ")
        kpis = payload.get("kpis") or {}
        line(f"Submissions pendientes: {kpis.get('submissions_pending', 0)}")
        line(f"Pagos pendientes: {kpis.get('payments_pending', 0)}")
        line(f"Alertas rojas abiertas: {kpis.get('red_alerts', 0)}")
        line(" ")

        line("Submissions (top 30):")
        for r in (payload.get("submissions") or [])[:30]:
            line(f"#{r.get('id')} {r.get('station_code')} {r.get('activity_title')} [{r.get('status')}] {r.get('created_at')}")
        line(" ")

        line("Payments (top 30):")
        for r in (payload.get("payments") or [])[:30]:
            line(f"#{r.get('id')} {r.get('station_code')} [{r.get('status')}] {r.get('created_at')}")
        line(" ")

        line("Alerts (top 30):")
        for r in (payload.get("alerts") or [])[:30]:
            line(f"#{r.get('id')} {r.get('station_code')} {r.get('severity')} [{r.get('status')}] {r.get('title')}")

        c.showPage(); c.save()
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name="inbox_export.pdf", mimetype="application/pdf")



    @app.get("/api/admin/missing-events")
    @login_required
    @role_required("admin")
    def api_admin_missing_events():
        """Calendar events without a valid submission (missing / reviewed / rejected) in a date range."""
        station_id = (request.args.get("station_id") or "").strip()
        date_from = (request.args.get("from") or "").strip()
        date_to = (request.args.get("to") or "").strip()

        today = datetime.date.today().isoformat()
        # defaults: current month
        if not date_to:
            date_to = today
        if not date_from:
            d = datetime.date.fromisoformat(date_to)
            date_from = d.replace(day=1).isoformat()

        conn = get_conn()
        cur = conn.cursor()

        # stations scope
        st_where = ""
        st_params = []
        if station_id:
            st_where = "WHERE id=?"
            st_params = [int(station_id)]
        cur.execute(f"SELECT id, code, name FROM stations {st_where} ORDER BY id", st_params)
        stations = [dict(r) for r in cur.fetchall()]

        # events in range (station-specific + global)
        cur.execute(
            "SELECT id, activity_id, title, start_date, repeat_kind, station_id "
            "FROM calendar_events WHERE brand=? AND date(start_date) >= date(?) AND date(start_date) <= date(?) "
            "ORDER BY date(start_date) ASC, id ASC",
            (get_brand(), date_from, date_to),
        )
        events = [dict(r) for r in cur.fetchall()]

        # preload activities titles
        cur.execute("SELECT id, title FROM activities WHERE brand=?", (get_brand(),))
        act_titles = {int(r["id"]): (r["title"] or "") for r in cur.fetchall()}

        # helper: latest submission for (event_id, station_id)
        def latest_submission(event_id: int, sid: int):
            cur.execute(
                "SELECT id, status, created_at FROM submissions "
                "WHERE brand=? AND event_id=? AND station_id=? ORDER BY id DESC LIMIT 1",
                (get_brand(), event_id, sid),
            )
            r = cur.fetchone()
            return dict(r) if r else None

        items = []
        overdue = 0
        pending = 0

        for ev in events:
            rk = (ev.get("repeat_kind") or "once").strip()
            base_title = act_titles.get(int(ev["activity_id"] or 0)) or ev.get("title") or "Actividad"
            ev_sid = ev.get("station_id")

            targets = []
            if ev_sid:
                targets = [s for s in stations if int(s["id"]) == int(ev_sid)]
            else:
                targets = stations  # global applies to all selected stations

            for st in targets:
                sid = int(st["id"])
                sub = latest_submission(int(ev["id"]), sid)
                status = (sub["status"] if sub else None)
                # valid completion: submitted or approved (admin will review later)
                valid = status in ("submitted", "approved")
                if valid:
                    continue

                is_overdue = (ev.get("start_date") or "") < today
                if is_overdue:
                    overdue += 1
                else:
                    pending += 1

                items.append(
                    {
                        "event_id": ev["id"],
                        "station_id": sid,
                        "station_code": st.get("code"),
                        "station_name": st.get("name"),
                        "date": ev.get("start_date"),
                        "repeat_kind": rk,
                        "title": base_title,
                        "last_status": status or "missing",
                        "last_submission_id": sub["id"] if sub else None,
                        "is_overdue": bool(is_overdue),
                    }
                )

        conn.close()
        # limit items to keep UI fast (still enough for admin)
        items = items[:400]
        return jsonify({"from": date_from, "to": date_to, "today": today, "overdue": overdue, "pending": pending, "items": items})


    @app.get("/api/admin/hub-metrics")
    @login_required
    @role_required("admin")
    def api_admin_hub_metrics():
        """Small summary metrics for the admin module hub UI."""
        today = datetime.date.today().isoformat()
        conn = get_conn()
        cur = conn.cursor()

        # Active stations (schema-safe)
        try:
            cur.execute("PRAGMA table_info(stations)")
            cols = {r["name"] for r in cur.fetchall()}
        except Exception:
            cols = set()

        if "is_active" in cols:
            cur.execute("SELECT COUNT(*) AS n FROM stations WHERE is_active=1")
        elif "monthly_status" in cols:
            cur.execute("SELECT COUNT(*) AS n FROM stations WHERE monthly_status='active'")
        else:
            # Fallback: count all stations
            cur.execute("SELECT COUNT(*) AS n FROM stations")

        stations_active = int(cur.fetchone()["n"])

        cur.execute("SELECT COUNT(*) AS n FROM alerts WHERE status='open'")
        alerts_open = int(cur.fetchone()["n"])

        cur.execute("SELECT COUNT(*) AS n FROM alerts WHERE status='open' AND severity='red'")
        alerts_red_open = int(cur.fetchone()["n"])

        # Today scheduled events
        cur.execute(
            """
            SELECT COUNT(*) AS n
            FROM calendar_events e
            WHERE date(e.start_date) = date(?)
            """,
            (today,),
        )
        events_today = int(cur.fetchone()["n"])

        # Today submissions
        cur.execute(
            """
            SELECT COUNT(*) AS n
            FROM submissions s
            WHERE date(s.created_at) = date(?)
            """,
            (today,),
        )
        submissions_today = int(cur.fetchone()["n"])

        # Pending reviews
        cur.execute("SELECT COUNT(*) AS n FROM submissions WHERE status IN ('submitted','reviewed')")
        reviews_pending = int(cur.fetchone()["n"])

        pending_today = max(events_today - submissions_today, 0)

        return jsonify(
            {
                "ok": True,
                "today": today,
                "stations_active": stations_active,
                "alerts_open": alerts_open,
                "alerts_red_open": alerts_red_open,
                "events_today": events_today,
                "submissions_today": submissions_today,
                "pending_today": pending_today,
                "reviews_pending": reviews_pending,
            }
        )
