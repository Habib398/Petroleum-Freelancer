from __future__ import annotations
import json, datetime, os
from flask import request, jsonify, session, redirect, render_template, send_from_directory, abort, current_app, Response
from werkzeug.security import generate_password_hash
from db import get_conn, verify_user, get_user
from services.brand import get_brand



ALLOWED_FUEL = ('magna','premium','diesel')

def register(app):
    ctx = app.extensions['ctx']
    login_required = ctx.login_required
    role_required = ctx.role_required

    def _pick_admin_station_id(explicit_station_id=None):
        """Resolve station_id for admin flows without assuming station #1 exists."""
        if explicit_station_id not in (None, "", 0, "0"):
            try:
                return int(explicit_station_id)
            except Exception:
                return None
        conn = get_conn(); cur = conn.cursor()
        cur.execute("SELECT id FROM stations WHERE brand=? ORDER BY id ASC LIMIT 1", (get_brand(),))
        row = cur.fetchone()
        conn.close()
        return int(row["id"]) if row else None


    @app.get("/api/reports/monthly.pdf")
    @login_required
    def report_monthly():
        me=ctx.get_me()
        if me and me.get("role")=="operador":
            return jsonify({"error":"forbidden"}),403
        if me and me.get("role")=="operador":
            return jsonify({"error":"forbidden"}),403
        if ctx.station_blocked(me) and me["role"]!="admin":
            return jsonify({"error":"station_blocked"}),403
        year=int(request.args.get("year") or datetime.date.today().year)
        month=int(request.args.get("month") or datetime.date.today().month)
        if me["role"]=="admin" and request.args.get("station_id"):
            station_id=int(request.args.get("station_id"))
        else:
            station_id = ctx.require_station(me) if me["role"]!="admin" else _pick_admin_station_id(request.args.get("station_id"))
        if station_id is None:
            return jsonify({"error":"no_station_available","message":"No hay estaciones registradas para generar el reporte."}), 400
        # build basic PDF with totals
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        import tempfile
        import math
        conn=get_conn(); cur=conn.cursor()
        cur.execute("SELECT * FROM stations WHERE brand=? AND id=?", (get_brand(), station_id,))
        st=cur.fetchone()
        cur.execute(
            "SELECT fuel_type, SUM(liters) as total FROM pipas WHERE brand=? AND station_id=? AND strftime('%Y',created_at)=? AND strftime('%m',created_at)=? GROUP BY fuel_type",
            (get_brand(), station_id,str(year),f"{month:02d}"),
        )
        fuel_tot={r["fuel_type"]: (r["total"] or 0) for r in cur.fetchall()}
        conn.close()

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        c = canvas.Canvas(tmp.name, pagesize=letter)
        w,h = letter
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, h-50, "COG WORK LOG - Reporte Mensual")
        c.setFont("Helvetica", 11)
        station_line = f"Estación: {st['name']} ({st['code']})" if st else f"Estación: {station_id}"
        c.drawString(40, h-75, station_line)
        c.drawString(40, h-90, f"Periodo: {year}-{month:02d}")
        y=h-130
        c.setFont("Helvetica-Bold", 11)
        c.drawString(40, y, "Litros recibidos:")
        y-=20
        c.setFont("Helvetica", 11)
        for ft in ALLOWED_FUEL:
            c.drawString(60, y, f"{ft.title()}: {fuel_tot.get(ft,0):,.2f} L")
            y-=16
        c.showPage(); c.save()
        tmp.close()
        ctx.log_action(me,"download_report_monthly","reports",f"{station_id}:{year}-{month:02d}")
        return send_from_directory(os.path.dirname(tmp.name), os.path.basename(tmp.name), as_attachment=True, download_name=f"reporte_mensual_{station_id}_{year}_{month:02d}.pdf")


    @app.get("/api/reports/annual.pdf")
    @login_required
    def report_annual():
        me=ctx.get_me()
        if ctx.station_blocked(me) and me["role"]!="admin":
            return jsonify({"error":"station_blocked"}),403
        year=int(request.args.get("year") or datetime.date.today().year)
        if me["role"]=="admin" and request.args.get("station_id"):
            station_id=int(request.args.get("station_id"))
        else:
            station_id = ctx.require_station(me) if me["role"]!="admin" else _pick_admin_station_id(request.args.get("station_id"))
        if station_id is None:
            return jsonify({"error":"no_station_available","message":"No hay estaciones registradas para generar el reporte."}), 400
        if not ctx.require_fiel_for_annual(me, station_id):
            return jsonify({"error":"fiel_required"}),403
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        import tempfile
        conn=get_conn(); cur=conn.cursor()
        cur.execute("SELECT * FROM stations WHERE brand=? AND id=?", (get_brand(), station_id,))
        st=cur.fetchone()
        cur.execute(
            "SELECT fuel_type, SUM(liters) as total FROM pipas WHERE brand=? AND station_id=? AND strftime('%Y',created_at)=? GROUP BY fuel_type",
            (get_brand(), station_id,str(year)),
        )
        fuel_tot={r["fuel_type"]: (r["total"] or 0) for r in cur.fetchall()}
        cur.execute("SELECT COUNT(*) as c FROM alerts WHERE brand=? AND station_id=? AND status='open'", (get_brand(), station_id,))
        alerts_open=cur.fetchone()["c"]
        cur.execute("SELECT COUNT(*) as c FROM maintenance WHERE brand=? AND station_id=? AND strftime('%Y',created_at)=?", (get_brand(), station_id,str(year)))
        maint_count=cur.fetchone()["c"]
        conn.close()

        tmp=tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        c=canvas.Canvas(tmp.name, pagesize=letter)
        w,h=letter
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40,h-50,"COG WORK LOG - Reporte Anual")
        c.setFont("Helvetica",11)
        station_line = f"Estación: {st['name']} ({st['code']})" if st else f"Estación: {station_id}"
        c.drawString(40,h-75,station_line)
        c.drawString(40,h-90,f"Año: {year}")
        y=h-130
        c.setFont("Helvetica-Bold",11)
        c.drawString(40,y,"Resumen anual")
        y-=20
        c.setFont("Helvetica",11)
        for ft in ALLOWED_FUEL:
            c.drawString(60,y,f"{ft.title()}: {fuel_tot.get(ft,0):,.2f} L")
            y-=16
        y-=10
        c.drawString(60,y,f"Alertas abiertas: {alerts_open}")
        y-=16
        c.drawString(60,y,f"Mantenimientos registrados: {maint_count}")
        c.showPage(); c.save()
        tmp.close()
        ctx.log_action(me,"download_report_annual","reports",f"{station_id}:{year}")
        return send_from_directory(os.path.dirname(tmp.name), os.path.basename(tmp.name), as_attachment=True, download_name=f"reporte_anual_{station_id}_{year}.pdf")


    @app.get("/api/reports/monthly.xlsx")
    @login_required
    def report_monthly_xlsx():
        me=ctx.get_me()
        if me and me.get("role")=="operador":
            return jsonify({"error":"forbidden"}),403
        if ctx.station_blocked(me) and me["role"]!="admin":
            return jsonify({"error":"station_blocked"}),403
        year=int(request.args.get("year") or datetime.date.today().year)
        month=int(request.args.get("month") or datetime.date.today().month)
        if me["role"]=="admin" and request.args.get("station_id"):
            station_id=int(request.args.get("station_id"))
        else:
            station_id = ctx.require_station(me) if me["role"]!="admin" else _pick_admin_station_id(request.args.get("station_id"))

        if station_id is None:
            return jsonify({"error":"no_station_available","message":"No hay estaciones registradas para generar el reporte."}), 400

        from openpyxl import Workbook
        import tempfile

        conn=get_conn(); cur=conn.cursor()
        cur.execute("SELECT * FROM stations WHERE brand=? AND id=?", (get_brand(), station_id,))
        st=cur.fetchone()
        cur.execute(
            "SELECT fuel_type, SUM(liters) as total FROM pipas WHERE brand=? AND station_id=? AND strftime('%Y',created_at)=? AND strftime('%m',created_at)=? GROUP BY fuel_type",
            (get_brand(), station_id,str(year),f"{month:02d}"),
        )
        fuel_tot={r["fuel_type"]: (r["total"] or 0) for r in cur.fetchall()}
        conn.close()

        wb=Workbook()
        ws=wb.active
        ws.title="Reporte Mensual"
        ws.append(["COG WORK LOG - Reporte Mensual"])
        ws.append(["Estación", f"{st['name']} ({st['code']})" if st else str(station_id)])
        ws.append(["Periodo", f"{year}-{month:02d}"])
        ws.append([])
        ws.append(["Combustible","Litros"])
        for ft in ALLOWED_FUEL:
            ws.append([ft, float(fuel_tot.get(ft,0) or 0)])
        ws.append([])
        ws.append(["Generado por", me.get("username","-")])
        ws.append(["Fecha", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")])

        tmp=tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)
        tmp.close()

        ctx.log_action(me,"download_report_monthly_xlsx","reports",f"{station_id}:{year}-{month:02d}")
        return send_from_directory(os.path.dirname(tmp.name), os.path.basename(tmp.name), as_attachment=True, download_name=f"reporte_mensual_{station_id}_{year}_{month:02d}.xlsx")


    @app.get("/api/reports/annual.xlsx")
    @login_required
    def report_annual_xlsx():
        me=ctx.get_me()
        if ctx.station_blocked(me) and me["role"]!="admin":
            return jsonify({"error":"station_blocked"}),403
        year=int(request.args.get("year") or datetime.date.today().year)
        if me["role"]=="admin" and request.args.get("station_id"):
            station_id=int(request.args.get("station_id"))
        else:
            station_id = ctx.require_station(me) if me["role"]!="admin" else _pick_admin_station_id(request.args.get("station_id"))
        if station_id is None:
            return jsonify({"error":"no_station_available","message":"No hay estaciones registradas para generar el reporte."}), 400
        if not ctx.require_fiel_for_annual(me, station_id):
            return jsonify({"error":"fiel_required"}),403

        from openpyxl import Workbook
        import tempfile

        conn=get_conn(); cur=conn.cursor()
        cur.execute("SELECT * FROM stations WHERE brand=? AND id=?", (get_brand(), station_id,))
        st=cur.fetchone()
        cur.execute(
            "SELECT fuel_type, SUM(liters) as total FROM pipas WHERE brand=? AND station_id=? AND strftime('%Y',created_at)=? GROUP BY fuel_type",
            (get_brand(), station_id,str(year)),
        )
        fuel_tot={r["fuel_type"]: (r["total"] or 0) for r in cur.fetchall()}
        cur.execute("SELECT COUNT(*) as c FROM alerts WHERE brand=? AND station_id=? AND status='open'", (get_brand(), station_id,))
        alerts_open=cur.fetchone()["c"]
        cur.execute("SELECT COUNT(*) as c FROM maintenance WHERE brand=? AND station_id=? AND strftime('%Y',created_at)=?", (get_brand(), station_id,str(year)))
        maint_count=cur.fetchone()["c"]
        conn.close()

        wb=Workbook()
        ws=wb.active
        ws.title="Reporte Anual"
        ws.append(["COG WORK LOG - Reporte Anual"])
        ws.append(["Estación", f"{st['name']} ({st['code']})" if st else str(station_id)])
        ws.append(["Año", year])
        ws.append([])
        ws.append(["Combustible","Litros"])
        for ft in ALLOWED_FUEL:
            ws.append([ft, float(fuel_tot.get(ft,0) or 0)])
        ws.append([])
        ws.append(["Alertas abiertas", int(alerts_open)])
        ws.append(["Mantenimientos registrados", int(maint_count)])
        ws.append([])
        ws.append(["Generado por", me.get("username","-")])
        ws.append(["Fecha", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")])

        tmp=tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)
        tmp.close()

        ctx.log_action(me,"download_report_annual_xlsx","reports",f"{station_id}:{year}")
        return send_from_directory(os.path.dirname(tmp.name), os.path.basename(tmp.name), as_attachment=True, download_name=f"reporte_anual_{station_id}_{year}.xlsx")





    

    @app.get("/api/reports/activities.pdf")
    @login_required
    def report_activities_monthly_pdf():
        """PDF mensual de actividades (bitácora del mes)."""
        me = ctx.get_me()
        if ctx.station_blocked(me) and me["role"]!="admin":
            return jsonify({"error":"station_blocked"}),403

        year = int(request.args.get("year") or datetime.date.today().year)
        month = int(request.args.get("month") or datetime.date.today().month)
        station_id = request.args.get("station_id")

        if me["role"] != "admin":
            station_id = str(ctx.require_station(me))
        else:
            station_id = station_id or _pick_admin_station_id()

        if station_id is None:
            return jsonify({"error":"no_station_available","message":"No hay estaciones registradas para generar el reporte."}), 400


        start = datetime.date(year, month, 1)
        end = datetime.date(year+1, 1, 1) if month==12 else datetime.date(year, month+1, 1)
        rng_from = start.isoformat()
        rng_to = (end - datetime.timedelta(days=1)).isoformat()

        conn=get_conn(); cur=conn.cursor()
        cur.execute("SELECT * FROM stations WHERE id=? AND brand=?", (int(station_id), get_brand()))
        st=cur.fetchone()

        cur.execute(
            "SELECT ce.start_date, ce.repeat_kind, a.title, a.description "
            "FROM calendar_events ce LEFT JOIN activities a ON a.id=ce.activity_id AND a.brand=ce.brand "
            "WHERE ce.brand=? AND ce.start_date>=? AND ce.start_date<=? AND (ce.station_id IS NULL OR ce.station_id=?) "
            "ORDER BY ce.start_date ASC",
            (get_brand(), rng_from, rng_to, int(station_id)),
        )
        items=[dict(r) for r in cur.fetchall()]
        conn.close()

        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import inch
        import tempfile
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        c = canvas.Canvas(tmp.name, pagesize=letter)
        w,h = letter

        brand_label = "Agenda" if get_brand()=="petroleum" else "Actividades"
        singular_label = "Agenda" if get_brand()=="petroleum" else "Actividad"
        filename_label = "agenda" if get_brand()=="petroleum" else "actividades"

        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, h-50, f"COG WORK LOG - {brand_label} del mes")
        c.setFont("Helvetica", 10)
        st_line = f"Estación: {st['name']} ({st['code']})" if st else f"Estación: {station_id}"
        c.drawString(40, h-70, st_line)
        c.drawString(40, h-84, f"Periodo: {year}-{month:02d}   Rango: {rng_from} a {rng_to}")
        c.drawString(40, h-98, f"Generado por: {me.get('username','-')} • Rol: {me.get('role','-')} • {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")

        # Legend
        c.setFont("Helvetica-Bold", 10)
        c.drawString(40, h-122, "Leyenda:")
        c.setFont("Helvetica", 10)
        c.drawString(110, h-122, "Azul=Diaria • Verde=Semanal • Amarillo=Mensual • Gris=Única")

        y = h-145
        c.setFont("Helvetica-Bold", 9)
        c.drawString(40, y, "Fecha")
        c.drawString(95, y, "Frecuencia")
        c.drawString(170, y, singular_label)
        c.drawString(390, y, "Indicaciones / instrucciones")
        y -= 12
        c.line(40, y, w-40, y)
        y -= 10

        def wrap(text, max_chars):
            text = (text or "").replace("\n"," ").strip()
            if not text:
                return [""]
            out=[]
            while len(text) > max_chars:
                cut = text.rfind(" ", 0, max_chars)
                if cut <= 0: cut = max_chars
                out.append(text[:cut].strip())
                text = text[cut:].strip()
            out.append(text)
            return out

        c.setFont("Helvetica", 9)
        for it in items:
            if y < 70:
                c.showPage()
                c.setFont("Helvetica", 9)
                y = h-60
            date = it.get("start_date","")
            rk = it.get("repeat_kind") or "once"
            title = it.get("title") or ("Agenda" if get_brand()=="petroleum" else "Actividad")
            desc_lines = wrap(it.get("description") or "", 55)
            title_lines = wrap(title, 28)

            # row height based on max lines
            lines_n = max(len(desc_lines), len(title_lines), 1)
            c.drawString(40, y, date)
            c.drawString(95, y, rk)
            for i,tl in enumerate(title_lines[:lines_n]):
                c.drawString(170, y - (i*10), tl)
            for i,dl in enumerate(desc_lines[:lines_n]):
                c.drawString(390, y - (i*10), dl)
            y -= (lines_n*10 + 6)

        c.showPage(); c.save()
        tmp.close()
        ctx.log_action(me,"download_report_activities_monthly","reports",f"{station_id}:{year}-{month:02d}")
        return send_from_directory(os.path.dirname(tmp.name), os.path.basename(tmp.name), as_attachment=True,
                                   download_name=f"{filename_label}_{station_id}_{year}_{month:02d}.pdf")


    @app.get("/api/reports/activities.ics")
    @login_required
    def report_activities_monthly_ics():
        """Export simple iCalendar (.ics) for the month."""
        me = ctx.get_me()
        if ctx.station_blocked(me) and me["role"] != "admin":
            return jsonify({"error": "station_blocked"}), 403

        year = int(request.args.get("year") or datetime.date.today().year)
        month = int(request.args.get("month") or datetime.date.today().month)
        station_id = request.args.get("station_id")
        if me["role"] != "admin":
            station_id = str(ctx.require_station(me))
        else:
            station_id = station_id or _pick_admin_station_id()

        if station_id is None:
            return jsonify({"error":"no_station_available","message":"No hay estaciones registradas para generar el reporte."}), 400

        start = datetime.date(year, month, 1)
        end = datetime.date(year + 1, 1, 1) if month == 12 else datetime.date(year, month + 1, 1)
        rng_from = start.isoformat()
        rng_to = (end - datetime.timedelta(days=1)).isoformat()

        conn = get_conn(); cur = conn.cursor()
        cur.execute(
            "SELECT ce.id, ce.start_date, ce.title, ce.repeat_kind FROM calendar_events ce "
            "WHERE ce.brand=? AND ce.start_date>=? AND ce.start_date<=? AND (ce.station_id IS NULL OR ce.station_id=?) "
            "ORDER BY ce.start_date ASC",
            (get_brand(), rng_from, rng_to, int(station_id)),
        )
        items = [dict(r) for r in cur.fetchall()]
        conn.close()

        def esc(s: str) -> str:
            return (s or "").replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,").replace("\n", "\\n")

        lines = [
            "BEGIN:VCALENDAR",
            "VERSION:2.0",
            "PRODID:-//COG Work Log//Activities//ES",
            "CALSCALE:GREGORIAN",
            "METHOD:PUBLISH",
        ]
        dtstamp = datetime.datetime.now(datetime.timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        filename_label = "agenda" if get_brand()=="petroleum" else "actividades"
        for it in items:
            d = (it.get("start_date") or "").replace("-", "")
            uid = f"cog-activity-{it.get('id')}@local"
            title = esc(it.get("title") or ("Agenda" if get_brand()=="petroleum" else "Actividad"))
            rk = esc(it.get("repeat_kind") or "once")
            lines += [
                "BEGIN:VEVENT",
                f"UID:{uid}",
                f"DTSTAMP:{dtstamp}",
                f"DTSTART;VALUE=DATE:{d}",
                f"SUMMARY:{title}",
                f"DESCRIPTION:Frecuencia: {rk}",
                "END:VEVENT",
            ]
        lines.append("END:VCALENDAR")
        blob = ("\r\n".join(lines) + "\r\n").encode("utf-8")

        ctx.log_action(me, "export_ics", "reports", f"{station_id}:{year}-{month:02d}")
        resp = Response(blob, mimetype="text/calendar; charset=utf-8")
        resp.headers["Content-Disposition"] = f'attachment; filename="{filename_label}_{station_id}_{year}_{month:02d}.ics"'
        return resp


    @app.get("/api/reports/bitacoras.pdf")
    @login_required
    def report_bitacoras_monthly_pdf():
        """PDF mensual de bitácoras (diaria/semanal/mensual)."""
        me = ctx.get_me()
        if ctx.station_blocked(me) and me["role"]!="admin":
            return jsonify({"error":"station_blocked"}),403

        year = int(request.args.get("year") or datetime.date.today().year)
        month = int(request.args.get("month") or datetime.date.today().month)
        station_id = request.args.get("station_id")
        if me["role"] != "admin":
            station_id = str(ctx.require_station(me))
        else:
            station_id = station_id or _pick_admin_station_id()

        if station_id is None:
            return jsonify({"error":"no_station_available","message":"No hay estaciones registradas para generar el reporte."}), 400

        start = datetime.date(year, month, 1)
        end = datetime.date(year+1, 1, 1) if month==12 else datetime.date(year, month+1, 1)
        rng_from = start.isoformat()
        rng_to = (end - datetime.timedelta(days=1)).isoformat()

        conn=get_conn(); cur=conn.cursor()
        cur.execute("SELECT * FROM stations WHERE id=? AND brand=?", (int(station_id), get_brand()))
        st=cur.fetchone()
        cur.execute(
            "SELECT b.ref_date, b.kind, b.notes, u.username as user_name "
            "FROM bitacoras b LEFT JOIN users u ON u.id=b.created_by "
            "WHERE b.station_id=? AND b.ref_date>=? AND b.ref_date<=? ORDER BY b.ref_date ASC",
            (int(station_id), rng_from, rng_to),
        )
        items=[dict(r) for r in cur.fetchall()]
        conn.close()

        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        import tempfile
        tmp=tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        c=canvas.Canvas(tmp.name, pagesize=letter)
        w,h=letter

        c.setFont("Helvetica-Bold",14)
        c.drawString(40,h-50,"COG WORK LOG - Bitácora del mes")
        c.setFont("Helvetica",10)
        st_line = f"Estación: {st['name']} ({st['code']})" if st else f"Estación: {station_id}"
        c.drawString(40,h-70,st_line)
        c.drawString(40,h-84,f"Periodo: {year}-{month:02d}   Rango: {rng_from} a {rng_to}")
        c.drawString(40,h-98,f"Generado por: {me.get('username','-')} • Rol: {me.get('role','-')} • {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")

        c.setFont("Helvetica-Bold",10)
        c.drawString(40,h-122,"Leyenda:")
        c.setFont("Helvetica",10)
        c.drawString(110,h-122,"daily=Diaria • weekly=Semanal • monthly=Mensual")

        y=h-145
        c.setFont("Helvetica-Bold",9)
        c.drawString(40,y,"Fecha")
        c.drawString(95,y,"Tipo")
        c.drawString(170,y,"Usuario")
        c.drawString(260,y,"Notas / indicaciones")
        y-=12
        c.line(40,y,w-40,y)
        y-=10
        c.setFont("Helvetica",9)

        def wrap(text,max_chars):
            text=(text or "").replace("\n"," ").strip()
            if not text: return [""]
            out=[]
            while len(text)>max_chars:
                cut=text.rfind(" ",0,max_chars)
                if cut<=0: cut=max_chars
                out.append(text[:cut].strip()); text=text[cut:].strip()
            out.append(text); return out

        for it in items:
            if y<70:
                c.showPage(); c.setFont("Helvetica",9); y=h-60
            date=it.get("ref_date","")
            kind=it.get("kind","")
            user=it.get("user_name","")
            notes_lines=wrap(it.get("notes") or "", 80)
            c.drawString(40,y,date)
            c.drawString(95,y,kind)
            c.drawString(170,y,user)
            for i,nl in enumerate(notes_lines):
                c.drawString(260,y-(i*10),nl)
            y -= (len(notes_lines)*10 + 6)

        c.showPage(); c.save()
        tmp.close()
        ctx.log_action(me,"download_report_bitacoras_monthly","reports",f"{station_id}:{year}-{month:02d}")
        return send_from_directory(os.path.dirname(tmp.name), os.path.basename(tmp.name), as_attachment=True,
                                   download_name=f"bitacora_{station_id}_{year}_{month:02d}.pdf")



    @app.get("/api/reports/consolidated.csv")
    @login_required
    @role_required("admin")
    def report_consolidated_csv():
        import io, csv
        brand = get_brand()
        conn=get_conn(); cur=conn.cursor()
        cur.execute("SELECT id, code, name FROM stations WHERE brand=? ORDER BY code ASC, id ASC", (brand,))
        stations=[dict(r) for r in cur.fetchall()]
        rows=[]
        for st in stations:
            sid=int(st['id'])
            cur.execute("SELECT COUNT(*) AS c FROM alerts WHERE brand=? AND station_id=? AND status='open'", (brand, sid))
            alerts_open=int(cur.fetchone()['c'] or 0)
            cur.execute("SELECT COUNT(*) AS c FROM maintenance WHERE brand=? AND station_id=?", (brand, sid))
            maint_total=int(cur.fetchone()['c'] or 0)
            cur.execute("SELECT COUNT(*) AS c FROM payments WHERE brand=? AND station_id=? AND status='pending'", (brand, sid))
            payments_pending=int(cur.fetchone()['c'] or 0)
            cur.execute("SELECT COUNT(*) AS c FROM doc_records WHERE brand=? AND station_id=? AND module='sasisopa'", (brand, sid))
            sasisopa_records=int(cur.fetchone()['c'] or 0)
            cur.execute("SELECT COUNT(*) AS c FROM doc_records WHERE brand=? AND station_id=? AND module='sgm'", (brand, sid))
            sgm_records=int(cur.fetchone()['c'] or 0)
            rows.append([st['code'], st['name'], alerts_open, maint_total, payments_pending, sasisopa_records, sgm_records])
        conn.close()
        out=io.StringIO(); w=csv.writer(out)
        w.writerow(["station_code","station_name","alerts_open","maintenance_total","payments_pending","sasisopa_records","sgm_records"])
        w.writerows(rows)
        ctx.log_action(ctx.get_me(), "download_report_consolidated_csv", "reports", brand, {"stations": len(rows)})
        return Response(out.getvalue().encode('utf-8'), mimetype='text/csv; charset=utf-8', headers={"Content-Disposition": f"attachment; filename=reporte_consolidado_{brand}.csv"})

    @app.get("/api/reports/consolidated.xlsx")
    @login_required
    @role_required("admin")
    def report_consolidated_xlsx():
        from io import BytesIO
        from openpyxl import Workbook
        brand = get_brand()
        conn=get_conn(); cur=conn.cursor()
        cur.execute("SELECT id, code, name FROM stations WHERE brand=? ORDER BY code ASC, id ASC", (brand,))
        stations=[dict(r) for r in cur.fetchall()]
        wb=Workbook(); ws=wb.active; ws.title='Consolidado'
        ws.append(['Reporte consolidado', brand])
        ws.append([])
        ws.append(['Código','Estación','Alertas abiertas','Mantenimientos','Pagos pendientes','Registros SASISOPA','Registros SGM'])
        for st in stations:
            sid=int(st['id'])
            cur.execute("SELECT COUNT(*) AS c FROM alerts WHERE brand=? AND station_id=? AND status='open'", (brand, sid)); a=int(cur.fetchone()['c'] or 0)
            cur.execute("SELECT COUNT(*) AS c FROM maintenance WHERE brand=? AND station_id=?", (brand, sid)); m=int(cur.fetchone()['c'] or 0)
            cur.execute("SELECT COUNT(*) AS c FROM payments WHERE brand=? AND station_id=? AND status='pending'", (brand, sid)); pnd=int(cur.fetchone()['c'] or 0)
            cur.execute("SELECT COUNT(*) AS c FROM doc_records WHERE brand=? AND station_id=? AND module='sasisopa'", (brand, sid)); sas=int(cur.fetchone()['c'] or 0)
            cur.execute("SELECT COUNT(*) AS c FROM doc_records WHERE brand=? AND station_id=? AND module='sgm'", (brand, sid)); sgm=int(cur.fetchone()['c'] or 0)
            ws.append([st['code'], st['name'], a, m, pnd, sas, sgm])
        conn.close()
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        ctx.log_action(ctx.get_me(), "download_report_consolidated_xlsx", "reports", brand, {"stations": len(stations)})
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment; filename=reporte_consolidado_{brand}.xlsx"},
        )

    @app.get("/api/reports/consolidated.pdf")
    @login_required
    @role_required("admin")
    def report_consolidated_pdf():
        from io import BytesIO
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        brand=get_brand()
        conn=get_conn(); cur=conn.cursor()
        cur.execute("SELECT id, code, name FROM stations WHERE brand=? ORDER BY code ASC, id ASC", (brand,))
        stations=[dict(r) for r in cur.fetchall()]
        buf = BytesIO()
        c=canvas.Canvas(buf, pagesize=letter); w,h=letter
        c.setFont('Helvetica-Bold', 14); c.drawString(40, h-50, 'COG WORK LOG - Reporte consolidado por estación')
        c.setFont('Helvetica', 10); c.drawString(40, h-66, f'Marca: {brand}')
        y=h-92
        c.setFont('Helvetica-Bold', 8)
        headers=[('Código',40),('Estación',90),('Alertas',260),('Mant.',315),('Pagos',360),('SAS',410),('SGM',450)]
        for label,x in headers: c.drawString(x,y,label)
        y-=10; c.line(40,y,w-40,y); y-=12; c.setFont('Helvetica',8)
        for st in stations:
            if y<60:
                c.showPage(); y=h-60; c.setFont('Helvetica',8)
            sid=int(st['id'])
            cur.execute("SELECT COUNT(*) AS c FROM alerts WHERE brand=? AND station_id=? AND status='open'", (brand, sid)); a=int(cur.fetchone()['c'] or 0)
            cur.execute("SELECT COUNT(*) AS c FROM maintenance WHERE brand=? AND station_id=?", (brand, sid)); m=int(cur.fetchone()['c'] or 0)
            cur.execute("SELECT COUNT(*) AS c FROM payments WHERE brand=? AND station_id=? AND status='pending'", (brand, sid)); pnd=int(cur.fetchone()['c'] or 0)
            cur.execute("SELECT COUNT(*) AS c FROM doc_records WHERE brand=? AND station_id=? AND module='sasisopa'", (brand, sid)); sas=int(cur.fetchone()['c'] or 0)
            cur.execute("SELECT COUNT(*) AS c FROM doc_records WHERE brand=? AND station_id=? AND module='sgm'", (brand, sid)); sgm=int(cur.fetchone()['c'] or 0)
            c.drawString(40,y,str(st['code'] or ''))
            c.drawString(90,y,(st['name'] or '')[:28])
            c.drawRightString(290,y,str(a)); c.drawRightString(340,y,str(m)); c.drawRightString(390,y,str(pnd)); c.drawRightString(430,y,str(sas)); c.drawRightString(470,y,str(sgm))
            y-=11
        conn.close(); c.showPage(); c.save()
        payload = buf.getvalue()
        ctx.log_action(ctx.get_me(), "download_report_consolidated_pdf", "reports", brand, {"stations": len(stations)})
        return Response(
            payload,
            mimetype='application/pdf',
            headers={"Content-Disposition": f"attachment; filename=reporte_consolidado_{brand}.pdf"},
        )

# Analytics + exports live in a dedicated module to keep this file smaller
    # and to ensure routes are always registered inside the app factory.
