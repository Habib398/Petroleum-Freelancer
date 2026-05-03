from __future__ import annotations

import datetime
import json
from pathlib import Path

from flask import Blueprint, request, jsonify, current_app
from werkzeug.utils import secure_filename

from db import get_conn
from services.brand import get_brand
from services.corrections import create_correction_task




# ---------------- upload validation ----------------
ALLOWED_EVIDENCE_EXT = {".pdf", ".png", ".jpg", ".jpeg"}
ALLOWED_IMPORT_EXT = {".csv", ".xlsx"}


def _validate_upload(fs) -> tuple[bool, str]:
    """Return (ok, error_code)."""
    if not fs or not getattr(fs, "filename", ""):
        return True, ""
    name = (fs.filename or "").strip()
    ext = (Path(name).suffix or "").lower()
    if ext not in ALLOWED_EVIDENCE_EXT:
        return False, "invalid_file_type"
    return True, ""


# ---------------- date helpers ----------------
def _end_of_year(date_str: str) -> str:
    try:
        y = int(date_str.split("-")[0])
    except Exception:
        y = datetime.date.today().year
    return f"{y}-12-31"


def _add_months(d: datetime.date, months: int) -> datetime.date:
    # add months, clamp day to end of resulting month
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    # last day of target month
    if m == 12:
        last = datetime.date(y + 1, 1, 1) - datetime.timedelta(days=1)
    else:
        last = datetime.date(y, m + 1, 1) - datetime.timedelta(days=1)
    day = min(d.day, last.day)
    return datetime.date(y, m, day)


def _generate_dates(start_date: str, repeat: str, until: str | None) -> list[datetime.date]:
    sd = datetime.date.fromisoformat(start_date)
    if not repeat or repeat == "once":
        return [sd]

    end = datetime.date.fromisoformat(until or _end_of_year(start_date))
    out: list[datetime.date] = []
    cur = sd
    while cur <= end:
        out.append(cur)

        if repeat == "daily":
            cur = cur + datetime.timedelta(days=1)
        elif repeat == "weekly":
            cur = cur + datetime.timedelta(days=7)
        elif repeat == "monthly":
            cur = _add_months(cur, 1)
        elif repeat == "bimonthly":
            cur = _add_months(cur, 2)
        elif repeat == "quarterly":
            cur = _add_months(cur, 3)
        elif repeat == "fourmonthly":
            cur = _add_months(cur, 4)
        elif repeat == "semiannual":
            cur = _add_months(cur, 6)
        elif repeat == "yearly":
            cur = _add_months(cur, 12)
        elif repeat == "fiveyearly":
            cur = _add_months(cur, 60)
        else:
            break

    return out


def _event_obj(row: dict) -> dict:
    """
    Convert DB row -> FullCalendar event object.
    Uses CSS classes (.rk-*) defined in theme-hme.css
    """
    rk = (row.get("repeat_kind") or "once").strip()
    title = row.get("activity_title") or row.get("title") or "Actividad"
    start = row.get("start_date")
    return {
        "id": row.get("id"),
        "title": title,
        "start": start,
        "allDay": True,
        "classNames": [f"rk-{rk}"],
        "extendedProps": {
            "repeat_kind": rk,
            "base_title": title,
            "activity_id": row.get("activity_id"),
            "station_id": row.get("station_id"),
            "submission_status": row.get("submission_status"),
        },
    }


def register(app):
    activities_bp = Blueprint("activities", __name__)
    ctx = app.extensions["ctx"]
    login_required = ctx.login_required
    role_required = ctx.role_required

    # ---------------- activity templates ----------------
    @activities_bp.get("/api/activities")
    @login_required
    @role_required("admin")
    def api_activities():
        conn = get_conn()
        cur = conn.cursor()
        brand = get_brand()
        cur.execute("SELECT * FROM activities WHERE brand=? ORDER BY id DESC",(brand,))
        activities = [dict(r) for r in cur.fetchall()]
        conn.close()
        return jsonify({"activities": activities})

    @activities_bp.post("/api/activities")
    @login_required
    @role_required("admin")
    def api_activity_create():
        me = ctx.get_me()
        payload = request.get_json(silent=True) or {}
        title = (payload.get("title") or "").strip()
        description = (payload.get("description") or "").strip()
        evidence_required = 1 if payload.get("evidence_required") else 0
        is_active = 1 if (payload.get("is_active", 1) in (1, True, "1", "true")) else 0

        if not title:
            return jsonify({"error": "missing_title"}), 400

        conn = get_conn()
        cur = conn.cursor()
        brand = get_brand()
        cur.execute(
            "INSERT INTO activities (brand, title, description, evidence_required, is_active, created_by) VALUES (?,?,?,?,?,?)",
            (brand, title, description, evidence_required, is_active, me["id"]),
        )
        conn.commit()
        aid = cur.lastrowid
        conn.close()
        ctx.log_action(me, "create_activity", "activities", str(aid))
        return jsonify({"ok": True, "id": aid})

    @activities_bp.put("/api/activities/<int:activity_id>")
    @login_required
    @role_required("admin")
    def api_activity_update(activity_id: int):
        me = ctx.get_me()
        payload = request.get_json(silent=True) or {}
        title = (payload.get("title") or "").strip()
        description = (payload.get("description") or "").strip()
        evidence_required = 1 if payload.get("evidence_required") else 0
        is_active = 1 if (payload.get("is_active", 1) in (1, True, "1", "true")) else 0

        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "UPDATE activities SET title=?, description=?, evidence_required=?, is_active=? WHERE id=? AND brand=?",
            (title, description, evidence_required, is_active, activity_id, get_brand()),
        )
        conn.commit()
        conn.close()
        ctx.log_action(me, "update_activity", "activities", str(activity_id))
        return jsonify({"ok": True})

    @activities_bp.delete("/api/activities/<int:activity_id>")
    @login_required
    @role_required("admin")
    def api_activity_delete(activity_id: int):
        """Soft delete: deactivate template to avoid breaking historical events."""
        me = ctx.get_me()
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("UPDATE activities SET is_active=0 WHERE id=? AND brand=?", (activity_id, get_brand()))
        conn.commit()
        conn.close()
        ctx.log_action(me, "delete_activity", "activities", str(activity_id))
        return jsonify({"ok": True})

    @activities_bp.post("/api/activity-templates")
    @login_required
    @role_required("admin")
    def api_activity_template_create():
        """Create an activity template + generate recurring calendar events (optionally by station)."""
        me = ctx.get_me()
        brand = get_brand()

        title = (request.form.get("title") or "").strip()
        description = (request.form.get("description") or "").strip()
        start_date = (request.form.get("start_date") or "").strip()
        repeat = (request.form.get("repeat") or "once").strip()
        until = (request.form.get("until") or "").strip() or None
        station_id_raw = (request.form.get("station_id") or "").strip()
        station_id = None if (not station_id_raw or station_id_raw in ("0", "null", "None")) else int(station_id_raw)

        if not title or not start_date:
            return jsonify({"error": "missing_fields"}), 400

        # Optional PDFs
        manual_pdf = request.files.get("manual_pdf")
        extra_pdf = request.files.get("extra_pdf")

        def save_pdf(f):
            if not f or not getattr(f, "filename", ""):
                return (None, None)
            fname = secure_filename(f.filename)
            if not fname.lower().endswith(".pdf"):
                return (None, None)
            limit = int(current_app.config.get("UPLOAD_LIMIT_DEFAULT_MB", 20))
            try:
                rel = ctx.save_upload_checked(f, "activities", allowed_ext={".pdf"}, allowed_magic={"pdf"}, limit_mb=limit)
            except Exception:
                return (None, None)
            return (rel, fname)

        manual_path, manual_name = save_pdf(manual_pdf)
        extra_path, extra_name = save_pdf(extra_pdf)

        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO activities (brand, title, description, evidence_required, is_active, created_by, manual_path, manual_name, extra_path, extra_name, recurrence, target_station_id) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (brand, title, description, 1, 1, me["id"], manual_path, manual_name, extra_path, extra_name, repeat, station_id),
        )
        activity_id = cur.lastrowid

        dates = _generate_dates(start_date, repeat, until)
        for d in dates:
            cur.execute(
                "INSERT INTO calendar_events (brand, activity_id, title, start_date, repeat_kind, station_id, created_by) VALUES (?,?,?,?,?,?,?)",
                (brand, activity_id, title, d.isoformat(), repeat, station_id, me["id"]),
            )
        conn.commit()
        conn.close()

        ctx.log_action(me, "create_activity_template", "activities", str(activity_id), {"repeat": repeat, "station_id": station_id, "events": len(dates)})
        # Activity template notifications are admin-only.
        extra = f" (Estación {station_id})" if station_id else " (Todas las estaciones)"
        ctx.notify_admins("Nueva actividad creada", f"{title}{extra}", "/mod/activities", station_id=station_id, exclude_user_id=me.get("id"), ntype="activity")
        return jsonify({"ok": True, "activity_id": activity_id, "events_created": len(dates)})


    @activities_bp.get("/api/import/activities/template")
    @login_required
    @role_required("admin")
    def api_import_activities_template():
        from flask import send_file
        from io import BytesIO
        base = Path(current_app.root_path) / "static" / "templates" / "activities_template.csv"
        data = base.read_bytes()
        return send_file(
            BytesIO(data),
            as_attachment=True,
            download_name="plantilla_actividades.csv",
            mimetype="text/csv; charset=utf-8",
        )

    # ---------------- import maintenance program ----------------
    @activities_bp.post("/api/import/maintenance-program")
    @login_required
    @role_required("admin")
    def api_import_maintenance_program():
        """Importar programa de mantenimiento preventivo anual y generar eventos."""
        me = ctx.get_me()
        brand = get_brand()
        payload = request.get_json(silent=True) or {}
        year = int(payload.get("year") or datetime.date.today().year)
        station_id = payload.get("station_id")  # None => todas
        if station_id is not None:
            station_id = int(station_id)

        conn = get_conn()
        cur = conn.cursor()

        # Avoid duplicates per year / station scope
        cur.execute(
            "SELECT COUNT(*) AS c FROM activities WHERE brand=? AND title LIKE ? AND (target_station_id IS ? OR target_station_id=?)",
            (brand, f"%[PM{year}]%", station_id, station_id),
        )
        if cur.fetchone()["c"] > 0:
            conn.close()
            return jsonify({"error": "already_imported"}), 400

        items = [
            ("daily", 1, "Limpieza general en áreas comunes, paredes, bardas, herrería, puertas/ventanas, señales y avisos."),
            ("daily", 2, "Limpieza general de sanitarios (empleados y públicos)."),
            ("daily", 3, "Limpieza en el exterior de dispensarios."),
            ("daily", 4, "Limpieza de registros y trampa de grasas para retirar aceites y sólidos gruesos."),

            ("monthly", 5, "Revisar que las luminarias de toda la estación estén funcionando correctamente."),
            ("monthly", 6, "Detección de fugas y derrames (anexar inventarios y tickets de alarmas de sensores)."),
            ("monthly", 7, "Lavado de pisos en áreas de despacho (agua y desengrasante)."),
            ("monthly", 8, "Limpieza en zona de almacenamiento (agua y desengrasante)."),
            ("monthly", 9, "Limpieza de registros y rejillas (retirar y lavar con agua y desengrasante)."),
            ("monthly", 10, "Inspección y limpieza de trampas de combustibles y de grasas (cuando se requiera); recolectar residuos/lodos en depósitos herméticos."),
            ("monthly", 11, "Drenado de tanques (sistema de control de inventario; anexar tickets)."),
            ("monthly", 12, "Verificación del funcionamiento del sistema de control de inventario (imprimir inventario)."),
            ("monthly", 13, "Limpieza e inspección de contenedores de bomba sumergible y accesorios/dispensario: sin fugas y sellado hermético."),
            ("monthly", 14, "Revisión de tinaco o cisterna."),
            ("monthly", 15, "Revisión de paros de emergencia e interruptores de emergencia."),
            ("monthly", 16, "Limpieza de contenedores en bocatoma de llenado: libres de combustible y herméticos."),

            ("quarterly", 17, "Retiro de residuos peligrosos generados (mantenimiento/limpieza) con empresas autorizadas."),
            ("quarterly", 18, "Revisión de flotadores en el sistema de medición."),

            ("fourmonthly", 19, "Pintura general: guarniciones, fachada de oficinas, señalamiento vertical y marcaje en pavimento."),
            ("fourmonthly", 20, "Limpieza de faldones y anuncio independiente."),

            ("semiannual", 21, "Revisión de interruptores/contactos/cajas de conexiones/sellos eléctricos/tableros, etc.; que tengan su correspondiente tapa."),

            ("yearly", 22, "Pruebas de hermeticidad en tanques y tuberías."),
            ("yearly", 23, "Recalibración de tanques de almacenamiento."),
            ("yearly", 24, "Revisión de continuidad eléctrica del sistema."),
            ("yearly", 25, "Mantenimiento de extintores (según NOM-002)."),

            ("fiveyearly", 26, "Limpieza interior de tanques de almacenamiento (autorización por escrito; indicar en bitácora fecha inicio y terminación)."),
        ]

        def dates_for(freq: str) -> list[datetime.date]:
            if freq == "daily":
                return _generate_dates(f"{year}-01-01", "daily", f"{year}-12-31")
            if freq == "monthly":
                return _generate_dates(f"{year}-01-01", "monthly", f"{year}-12-31")
            if freq == "quarterly":
                return _generate_dates(f"{year}-01-01", "quarterly", f"{year}-12-31")
            if freq == "fourmonthly":
                return _generate_dates(f"{year}-01-01", "fourmonthly", f"{year}-12-31")
            if freq == "semiannual":
                return [datetime.date(year, 1, 1), datetime.date(year, 7, 1)]
            if freq == "yearly":
                return [datetime.date(year, 1, 1)]
            if freq == "fiveyearly":
                return [datetime.date(year, 1, 1)]
            return [datetime.date(year, 1, 1)]

        created = 0
        for freq, num, desc in items:
            title = f"[PM{year}] #{num:02d} • {desc[:60]}..."
            cur.execute(
                "INSERT INTO activities (brand, title, description, evidence_required, is_active, created_by, recurrence, target_station_id) VALUES (?,?,?,?,?,?,?,?)",
                (brand, title, desc, 1, 1, me["id"], freq, station_id),
            )
            aid = cur.lastrowid
            for d in dates_for(freq):
                cur.execute(
                    "INSERT INTO calendar_events (brand, activity_id, title, start_date, repeat_kind, station_id, created_by) VALUES (?,?,?,?,?,?,?)",
                    (brand, aid, title, d.isoformat(), freq, station_id, me["id"]),
                )
                created += 1

        conn.commit()
        conn.close()
        ctx.log_action(me, "import_maintenance_program", "activities", str(year), {"station_id": station_id, "events": created})
        # Maintenance-program imports are admin-only notifications.
        extra = f" (Estación {station_id})" if station_id else " (Todas las estaciones)"
        ctx.notify_admins("Programa de mantenimiento importado", f"Año {year}{extra}", "/mod/activities", station_id=station_id, exclude_user_id=me.get("id"), ntype="activity")
        return jsonify({"ok": True, "events_created": created})

    # ---------------- import activities (CSV/XLSX) ----------------
    @activities_bp.post("/api/import/activities")
    @login_required
    @role_required("admin")
    def api_import_activities_file():
        """Importar actividades desde un archivo CSV/XLSX.

        Columnas soportadas (cabeceras, no sensibles a mayúsculas/espacios):
        - title / titulo (requerido)
        - start_date / fecha (YYYY-MM-DD) (requerido)
        - repeat / frecuencia (once,daily,weekly,monthly,quarterly,fourmonthly,semiannual,yearly)
        - until / hasta (YYYY-MM-DD opcional)
        - station_code / station_id (opcional). Vacío => todas.
        - description / descripcion (opcional)

        Crea plantilla en activities + genera calendar_events.
        """
        me = ctx.get_me()
        brand = get_brand()

        f = request.files.get("file")
        if not f or not getattr(f, "filename", ""):
            return jsonify({"error": "missing_file"}), 400
        fname = secure_filename(f.filename)
        ext = (Path(fname).suffix or "").lower()
        if ext not in ALLOWED_IMPORT_EXT:
            return jsonify({"error": "invalid_file_type", "allowed": sorted(ALLOWED_IMPORT_EXT)}), 400

        def norm(s: str) -> str:
            return (s or "").strip().lower().replace(" ", "_")

        rows: list[dict] = []
        try:
            if ext == ".csv":
                import csv
                content = f.stream.read().decode("utf-8", errors="ignore").splitlines()
                reader = csv.DictReader(content)
                for r in reader:
                    rows.append({norm(k): (v.strip() if isinstance(v, str) else v) for k, v in (r or {}).items()})
            else:
                from openpyxl import load_workbook
                wb = load_workbook(f, data_only=True)
                ws = wb.active
                headers = [norm(str(c.value or "")) for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for rr in ws.iter_rows(min_row=2):
                    obj = {}
                    for i, c in enumerate(rr):
                        if i >= len(headers):
                            continue
                        key = headers[i]
                        if not key:
                            continue
                        val = c.value
                        if isinstance(val, datetime.datetime):
                            val = val.date().isoformat()
                        elif isinstance(val, datetime.date):
                            val = val.isoformat()
                        elif val is None:
                            val = ""
                        else:
                            val = str(val).strip()
                        obj[key] = val
                    if any(v for v in obj.values()):
                        rows.append(obj)
        except Exception as e:
            return jsonify({"error": "parse_failed", "details": str(e)}), 400

        if not rows:
            return jsonify({"error": "empty_file"}), 400

        # Helper: resolve station
        conn = get_conn(); cur = conn.cursor()
        station_code_to_id = {}
        cur.execute("SELECT id, code FROM stations WHERE brand=?", (brand,))
        for r in cur.fetchall():
            station_code_to_id[(r["code"] or "").strip().upper()] = int(r["id"])

        def get_station_id(obj: dict):
            sc = (obj.get("station_code") or obj.get("codigo") or "").strip().upper()
            if sc:
                return station_code_to_id.get(sc)
            sid = (obj.get("station_id") or "").strip()
            if sid:
                try:
                    return int(float(sid))
                except Exception:
                    return None
            return None

        created_templates = 0
        created_events = 0
        errors: list[dict] = []

        for idx, r in enumerate(rows, start=2):
            title = (r.get("title") or r.get("titulo") or "").strip()
            start_date = (r.get("start_date") or r.get("fecha") or "").strip()
            repeat = (r.get("repeat") or r.get("frecuencia") or r.get("recurrence") or "once").strip().lower()
            until = (r.get("until") or r.get("hasta") or "").strip() or None
            desc = (r.get("description") or r.get("descripcion") or "").strip()
            station_id = get_station_id(r)

            if not title or not start_date:
                errors.append({"row": idx, "error": "missing_title_or_date"})
                continue
            try:
                # validate date
                datetime.date.fromisoformat(start_date[:10])
                if until:
                    datetime.date.fromisoformat(until[:10])
            except Exception:
                errors.append({"row": idx, "error": "bad_date"})
                continue

            allowed_repeat = {"once","daily","weekly","monthly","bimonthly","quarterly","fourmonthly","semiannual","yearly","fiveyearly"}
            if repeat not in allowed_repeat:
                repeat = "once"

            # Create template
            cur.execute(
                "INSERT INTO activities (brand, title, description, evidence_required, is_active, created_by, recurrence, target_station_id) VALUES (?,?,?,?,?,?,?,?)",
                (brand, title, desc or None, 1, 1, me["id"], repeat, station_id),
            )
            aid = cur.lastrowid
            created_templates += 1

            # Generate events
            dates = _generate_dates(start_date[:10], repeat, until[:10] if until else None)
            # Guardrail: avoid huge daily imports without until
            if repeat == "daily" and not until and len(dates) > 366:
                dates = dates[:366]
            for d in dates:
                cur.execute(
                    "INSERT INTO calendar_events (brand, activity_id, title, start_date, repeat_kind, station_id, created_by) VALUES (?,?,?,?,?,?,?)",
                    (brand, aid, title, d.isoformat(), repeat, station_id, me["id"]),
                )
                created_events += 1

        conn.commit(); conn.close()
        ctx.log_action(me, "import_activities", "activities", ext, {"templates": created_templates, "events": created_events, "errors": len(errors)})
        ctx.notify_admins("Actividades importadas", f"Plantillas: {created_templates} · Eventos: {created_events}", "/mod/activities", exclude_user_id=me.get("id"), ntype="activity")
        return jsonify({"ok": True, "templates_created": created_templates, "events_created": created_events, "errors": errors[:50]})

    # ---------------- calendar events ----------------
    @activities_bp.get("/api/calendar/events")
    @login_required
    def api_calendar_events():
        me = ctx.get_me()
        start = request.args.get("start", "")
        end = request.args.get("end", "")

        conn = get_conn()
        cur = conn.cursor()

        brand = get_brand()
        if me["role"] == "admin":
            cur.execute(
                "SELECT e.*, a.title AS activity_title FROM calendar_events e "
                "LEFT JOIN activities a ON a.id=e.activity_id AND a.brand=e.brand AND a.brand=e.brand "
                "WHERE e.brand=? AND (?='' OR e.start_date>=?) AND (?='' OR e.start_date<=?) "
                "ORDER BY e.start_date ASC",
                (brand, start, start, end, end),
            )
            rows = [dict(r) for r in cur.fetchall()]
            conn.close()
            return jsonify([_event_obj(r) for r in rows])

        # Non-admin: station scope + submission status join
        sid = ctx.require_station(me)
        cur.execute(
            "SELECT e.*, a.title AS activity_title FROM calendar_events e "
            "LEFT JOIN activities a ON a.id=e.activity_id AND a.brand=e.brand AND a.brand=e.brand "
            "WHERE e.brand=? AND (e.station_id IS NULL OR e.station_id=?) "
            "AND (?='' OR e.start_date>=?) AND (?='' OR e.start_date<=?) "
            "ORDER BY e.start_date ASC",
            (brand, sid, start, start, end, end),
        )
        rows = [dict(r) for r in cur.fetchall()]
        ev_ids = [r["id"] for r in rows]
        status_map = {}
        if ev_ids:
            q = ",".join("?" * len(ev_ids))
            cur.execute(f"SELECT event_id, status FROM submissions WHERE brand=? AND station_id=? AND event_id IN ({q})", (brand, sid, *ev_ids))
            status_map = {rr["event_id"]: rr["status"] for rr in cur.fetchall()}

        for r in rows:
            r["submission_status"] = status_map.get(r["id"])
        conn.close()
        return jsonify([_event_obj(r) for r in rows])

    @activities_bp.get("/api/calendar/events/<int:event_id>")
    @login_required
    def api_calendar_event_detail(event_id: int):
        me = ctx.get_me()
        conn = get_conn()
        cur = conn.cursor()

        cur.execute(
            "SELECT e.*, a.title AS activity_title, a.description AS activity_description, a.manual_path, a.manual_name, a.extra_path, a.extra_name, a.evidence_required "
            "FROM calendar_events e LEFT JOIN activities a ON a.id=e.activity_id AND a.brand=e.brand AND a.brand=e.brand WHERE e.brand=? AND e.id=?",
            (get_brand(), event_id),
        )
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "not_found"}), 404
        ev = dict(row)

        # enforce station visibility for non-admin
        if me["role"] != "admin":
            sid = ctx.require_station(me)
            if ev.get("station_id") is not None and int(ev["station_id"]) != sid:
                conn.close()
                return jsonify({"error": "forbidden"}), 403

            cur.execute(
                "SELECT * FROM submissions WHERE brand=? AND station_id=? AND event_id=? ORDER BY id DESC LIMIT 1",
                (get_brand(), sid, event_id),
            )
            sub = cur.fetchone()
            ev["submission"] = dict(sub) if sub else None
        else:
            ev["submission"] = None

        conn.close()
        return jsonify({"event": ev})

    @activities_bp.post("/api/calendar/events")
    @login_required
    @role_required("admin")
    def api_calendar_event_create():
        me = ctx.get_me()
        brand = get_brand()
        payload = request.get_json(silent=True) or {}
        title = (payload.get("title") or "").strip()
        start_date = (payload.get("start_date") or "").strip()
        repeat_kind = (payload.get("repeat_kind") or "once").strip()
        station_id = payload.get("station_id")
        activity_id = payload.get("activity_id")

        if station_id in ("", "null", None):
            station_id = None
        else:
            station_id = int(station_id)

        if not title or not start_date:
            return jsonify({"error": "missing_fields"}), 400

        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO calendar_events (brand, activity_id, title, start_date, repeat_kind, station_id, created_by) VALUES (?,?,?,?,?,?,?)",
            (brand, activity_id, title, start_date, repeat_kind, station_id, me["id"]),
        )
        conn.commit()
        eid = cur.lastrowid
        conn.close()
        ctx.log_action(me, "create_calendar_event", "calendar_events", str(eid))
        return jsonify({"ok": True, "id": eid})

    @activities_bp.put("/api/calendar/events/<int:event_id>")
    @login_required
    @role_required("admin")
    def api_calendar_event_update(event_id: int):
        me = ctx.get_me()
        payload = request.get_json(silent=True) or {}
        title = (payload.get("title") or "").strip()
        start_date = (payload.get("start_date") or "").strip()
        repeat_kind = (payload.get("repeat_kind") or "once").strip()
        station_id = payload.get("station_id")
        if station_id in ("", "null", None):
            station_id = None
        else:
            station_id = int(station_id)

        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "UPDATE calendar_events SET title=?, start_date=?, repeat_kind=?, station_id=? WHERE id=? AND brand=?",
            (title, start_date, repeat_kind, station_id, event_id, get_brand()),
        )
        conn.commit()
        conn.close()
        ctx.log_action(me, "update_calendar_event", "calendar_events", str(event_id))
        return jsonify({"ok": True})

    
    @activities_bp.post("/api/calendar/events/<int:event_id>/move")
    @login_required
    @role_required("admin")
    def api_calendar_event_move(event_id: int):
        """Move a calendar event (single) or shift the remaining events of the same template (series)."""
        me = ctx.get_me()
        payload = request.get_json(silent=True) or {}
        new_date = (payload.get("new_date") or "").strip()
        old_date = (payload.get("old_date") or "").strip()
        scope = (payload.get("scope") or "single").strip()

        if not new_date:
            return jsonify({"error": "missing_new_date"}), 400

        # validate date format YYYY-MM-DD
        try:
            new_dt = datetime.date.fromisoformat(new_date[:10])
        except Exception:
            return jsonify({"error": "bad_new_date"}), 400

        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, activity_id, station_id, repeat_kind, start_date, title, brand FROM calendar_events WHERE id=? AND brand=?", (event_id, get_brand()))
        row = cur.fetchone()
        if not row:
            conn.close()
            return jsonify({"error": "not_found"}), 404
        ev = dict(row)

        # determine old date
        base_old = old_date or (ev.get("start_date") or "")
        try:
            old_dt = datetime.date.fromisoformat(base_old[:10])
        except Exception:
            old_dt = datetime.date.fromisoformat((ev.get("start_date") or "")[:10])

        delta = (new_dt - old_dt).days

        if scope == "series" and ev.get("activity_id"):
            # shift all future events of the same template by delta days
            aid = int(ev["activity_id"])
            rk = (ev.get("repeat_kind") or "").strip()
            sid = ev.get("station_id")
            # update only events on/after the old date
            if sid is None:
                cur.execute(
                    "SELECT id, start_date FROM calendar_events WHERE brand=? AND activity_id=? AND (repeat_kind=? OR ?='') AND station_id IS NULL AND start_date>=?",
                    (get_brand(), aid, rk, rk, old_dt.isoformat()),
                )
            else:
                cur.execute(
                    "SELECT id, start_date FROM calendar_events WHERE brand=? AND activity_id=? AND (repeat_kind=? OR ?='') AND station_id=? AND start_date>=?",
                    (get_brand(), aid, rk, rk, int(sid), old_dt.isoformat()),
                )
            rows = cur.fetchall()
            for rr in rows:
                sd = rr["start_date"][:10]
                try:
                    d0 = datetime.date.fromisoformat(sd)
                except Exception:
                    continue
                d1 = d0 + datetime.timedelta(days=delta)
                cur.execute("UPDATE calendar_events SET start_date=? WHERE id=? AND brand=?", (d1.isoformat(), rr["id"], get_brand()))
            conn.commit()
            conn.close()
            ctx.log_action(me, "move_calendar_event_series", "calendar_events", str(event_id), {"delta_days": delta, "count": len(rows)})
            return jsonify({"ok": True, "scope": "series", "moved": len(rows)})

        # default: single
        cur.execute("UPDATE calendar_events SET start_date=? WHERE id=? AND brand=?", (new_dt.isoformat(), event_id, get_brand()))
        conn.commit()
        conn.close()
        ctx.log_action(me, "move_calendar_event", "calendar_events", str(event_id), {"delta_days": delta})
        return jsonify({"ok": True, "scope": "single"})


    @activities_bp.delete("/api/calendar/events/<int:event_id>")
    @login_required
    @role_required("admin")
    def api_calendar_event_delete(event_id: int):
        me = ctx.get_me()
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("DELETE FROM submissions WHERE event_id=? AND brand=?", (event_id, get_brand()))
        cur.execute("DELETE FROM calendar_events WHERE id=? AND brand=?", (event_id, get_brand()))
        conn.commit()
        conn.close()
        ctx.log_action(me, "delete_calendar_event", "calendar_events", str(event_id))
        return jsonify({"ok": True})

    # ---------------- submissions ----------------
    @activities_bp.get("/api/submissions")
    @login_required
    def api_submissions_list():
        """List submissions for current station (non-admin) or all (admin) - used for export/debug."""
        me = ctx.get_me()
        conn = get_conn()
        cur = conn.cursor()

        brand = get_brand()
        if me["role"] == "admin":
            cur.execute("SELECT * FROM submissions WHERE brand=? ORDER BY id DESC LIMIT 500", (brand,))
            rows = [dict(r) for r in cur.fetchall()]
            conn.close()
            return jsonify({"submissions": rows})

        sid = ctx.require_station(me)
        cur.execute("SELECT * FROM submissions WHERE brand=? AND station_id=? ORDER BY id DESC LIMIT 200", (brand, sid,))
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return jsonify({"submissions": rows})

    @activities_bp.post("/api/submissions")
    @login_required
    def api_submission_create():
        me = ctx.get_me()

        # enforce station permissions
        if me["role"] == "admin":
            return jsonify({"error": "admin_cannot_submit"}), 400

        if ctx.station_blocked(me):
            return jsonify({"error": "station_blocked"}), 403

        event_id = int(request.form.get("event_id") or 0)
        activity_id = int(request.form.get("activity_id") or 0)
        notes = (request.form.get("notes") or "").strip()
        evidence = request.files.get("evidence")
        signature_name = (request.form.get("signature_name") or "").strip() or None
        signature_ip = (request.headers.get("X-Forwarded-For") or request.remote_addr or "").split(",")[0].strip() or None

        if not event_id:
            return jsonify({"error": "missing_event"}), 400

        sid = ctx.require_station(me)

        # Optional admin-configured lock date for captures/evidence.
        try:
            lock_key = f"activity_lock_date:{get_brand()}"
            conn0 = get_conn(); row0 = conn0.execute("SELECT value FROM system_state WHERE key=?", (lock_key,)).fetchone(); conn0.close()
            lock_date = (row0["value"] if row0 else "") or ""
            if lock_date:
                import datetime as _dt
                if _dt.date.today() > _dt.date.fromisoformat(lock_date[:10]):
                    return jsonify({"error": "activity_locked_by_deadline", "message": f"La captura quedó cerrada desde {lock_date[:10]}"}), 403
        except Exception:
            pass

        ok_up, err_code = _validate_upload(evidence)
        if not ok_up:
            return jsonify({"error": err_code}), 400

        # Determine upload limit: annual evidence can be bigger
        is_annual = False
        try:
            conn = get_conn(); cur = conn.cursor()
            # Prefer explicit activity_id, else pull from event
            aid = activity_id
            if not aid and event_id:
                cur.execute("SELECT activity_id FROM calendar_events WHERE id=? AND brand=?", (event_id, get_brand()))
                r = cur.fetchone()
                aid = int(r["activity_id"] or 0) if r else 0
            if aid:
                cur.execute("SELECT recurrence FROM activities WHERE id=? AND brand=?", (aid, get_brand()))
                r2 = cur.fetchone()
                rec = (r2["recurrence"] if r2 else "")
                is_annual = (rec or "").lower() in {"annual", "yearly", "anual"}
            conn.close()
        except Exception:
            try:
                conn.close()
            except Exception:
                pass
            is_annual = False

        limit = int(current_app.config.get("UPLOAD_LIMIT_ANNUAL_MB" if is_annual else "UPLOAD_LIMIT_DEFAULT_MB", 120 if is_annual else 20))
        allowed_magic = {"pdf", "png", "jpg"}

        # save evidence file (optional)
        evidence_path = None
        if evidence and getattr(evidence, "filename", ""):
            try:
                evidence_path = ctx.save_upload_checked(
                    evidence,
                    subdir=f"submissions/{sid}",
                    allowed_ext=ALLOWED_EVIDENCE_EXT,
                    allowed_magic=allowed_magic,
                    limit_mb=limit,
                )
            except ValueError:
                return jsonify({"error": "invalid_file_type"}), 400

        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO submissions (brand, event_id, activity_id, station_id, user_id, notes, evidence_path, status, created_at) "
            "VALUES (?,?,?,?,?,?,?, 'submitted', CURRENT_TIMESTAMP)",
            (get_brand(), event_id, activity_id or None, sid, me["id"], notes, evidence_path),
        )
        # signature columns are optional migrations; update if present
        try:
            cur.execute(
                "UPDATE submissions SET signature_name=?, signature_ip=?, signature_at=CURRENT_TIMESTAMP, signature_role=? WHERE id=?",
                (signature_name, signature_ip, me.get("role"), cur.lastrowid),
            )
        except Exception:
            pass
        conn.commit()
        sub_id = cur.lastrowid
        conn.close()

        ctx.log_action(me, "create_submission", "submissions", str(sub_id), {"event_id": event_id})
        ctx.sign_entity(me, "submission", str(sub_id), "submitted", {"event_id": event_id, "station_id": sid, "signature_name": signature_name})
        # Activity submissions must be visible to admins only.
        ctx.notify_admins(
            "Nueva entrega",
            f"Evento #{event_id} (Estación {sid})",
            "/admin/inbox",
            station_id=sid,
            exclude_user_id=me.get("id"),
            ntype="submission",
        )
        return jsonify({"ok": True, "id": sub_id})

    @activities_bp.post("/api/submissions/<int:submission_id>/review")
    @login_required
    @role_required("admin", "jefe_estacion")
    def api_submission_review(submission_id: int):
        """Review workflow: set status (reviewed/approved/rejected) + notes/score.

        - admin can review any station.
        - jefe_estacion can only review their own station (or group).
        """
        me = ctx.get_me()
        payload = request.get_json(silent=True) or {}
        status = (payload.get("status") or "").strip().lower()
        review_notes = (payload.get("review_notes") or "").strip()
        score = payload.get("score")

        if status not in {"reviewed", "approved", "rejected"}:
            return jsonify({"error": "invalid_status"}), 400

        if score is not None:
            try:
                score = int(score)
            except Exception:
                score = None

        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT * FROM submissions WHERE id=? AND brand=?", (submission_id, get_brand()))
        sub = cur.fetchone()
        if not sub:
            conn.close()
            return jsonify({"error": "not_found"}), 404

        # jefe_estacion scope enforcement
        if me["role"] == "jefe_estacion":
            sid_me = ctx.require_station(me)
            if int(sub["station_id"] or 0) != int(sid_me):
                conn.close()
                return jsonify({"error": "forbidden_station"}), 403

        cur.execute(
            "UPDATE submissions SET status=?, review_notes=?, score=?, reviewed_by=?, reviewed_at=CURRENT_TIMESTAMP WHERE id=? AND brand=?",
            (status, review_notes or None, score, me["id"], submission_id, get_brand()),
        )
        conn.commit()

        # Notify the original user (operator) with a clear workflow message
        user_id = sub["user_id"]
        station_id = sub["station_id"]
        title = "Entrega aprobada" if status == "approved" else "Entrega rechazada" if status == "rejected" else "Entrega en revisión"
        body = review_notes or f"Entrega #{submission_id}"
        if user_id:
            ctx.notify(int(user_id), int(station_id) if station_id else None, title, body, f"/mod/activities/event/{sub['event_id']}", ntype="submission")

        # Rejected evidence must notify admins only.
        if status == "rejected":
            ctx.notify_admins(
                "Evidencia rechazada",
                f"Entrega #{submission_id}: {body}",
                "/admin/inbox",
                station_id=int(station_id),
                exclude_user_id=me.get("id"),
                ntype="submission",
            )
            try:
                create_correction_task(
                    ctx, me, brand=get_brand(), title=f"Corregir entrega #{submission_id}", description=body,
                    station_id=station_id, module="activities", related_entity="submission", related_entity_id=str(submission_id),
                    assigned_to=user_id, source_status="rejected", priority="high", due_days=3,
                )
            except Exception:
                pass

        ctx.log_action(me, "review_submission", "submissions", str(submission_id), {"status": status, "score": score})
        ctx.sign_entity(me, "submission", str(submission_id), f"review_{status}", {"status": status, "score": score, "review_notes": review_notes})
        conn.close()
        return jsonify({"ok": True})

    # ---------------- progress (dashboard) ----------------
    @activities_bp.get("/api/my/activity-progress")
    @login_required
    def api_my_activity_progress():
        """Progress of calendar activities for the current user's station (daily / monthly / yearly)."""
        me = ctx.get_me()
        if me["role"] == "admin":
            return jsonify({"error": "admin_no_station"}), 400

        sid = ctx.require_station(me)
        today = datetime.date.today()
        today_str = today.isoformat()
        month_start = today.replace(day=1).isoformat()
        year_start = today.replace(month=1, day=1).isoformat()

        conn = get_conn()
        cur = conn.cursor()

        def calc(range_start: str, range_end: str | None):
            if range_end:
                cur.execute(
                    "SELECT COUNT(*) AS c FROM calendar_events WHERE brand=? AND (station_id IS NULL OR station_id=?) AND start_date>=? AND start_date<=?",
                    (get_brand(), sid, range_start, range_end),
                )
            else:
                cur.execute(
                    "SELECT COUNT(*) AS c FROM calendar_events WHERE brand=? AND (station_id IS NULL OR station_id=?) AND start_date>=?",
                    (get_brand(), sid, range_start),
                )
            total = int(cur.fetchone()["c"])

            if range_end:
                cur.execute(
                    "SELECT COUNT(DISTINCT s.event_id) AS c FROM submissions s "
                    "WHERE s.brand=? AND s.station_id=? AND s.status IN ('submitted','approved') "
                    "AND s.event_id IN (SELECT id FROM calendar_events WHERE brand=? AND (station_id IS NULL OR station_id=?) AND start_date>=? AND start_date<=?)",
                    (get_brand(), sid, get_brand(), sid, range_start, range_end),
                )
            else:
                cur.execute(
                    "SELECT COUNT(DISTINCT s.event_id) AS c FROM submissions s "
                    "WHERE s.brand=? AND s.station_id=? AND s.status IN ('submitted','approved') "
                    "AND s.event_id IN (SELECT id FROM calendar_events WHERE brand=? AND (station_id IS NULL OR station_id=?) AND start_date>=?)",
                    (get_brand(), sid, get_brand(), sid, range_start),
                )
            done = int(cur.fetchone()["c"])
            pct = int(round((done / total) * 100)) if total else 0
            return {"done": done, "total": total, "pct": pct}

        daily = calc(today_str, today_str)
        monthly = calc(month_start, today_str)
        yearly = calc(year_start, today_str)

        conn.close()
        return jsonify({"today": today_str, "daily": daily, "monthly": monthly, "yearly": yearly})

    @activities_bp.get("/api/station/activity-progress")
    @login_required
    def api_station_activity_progress():
        """Progress of calendar activities for a station (admin/jefe_estacion/otros con permiso)."""
        me = ctx.get_me()
        station_id = request.args.get("station_id", type=int)

        conn = get_conn(); cur = conn.cursor()

        def allowed_station_ids():
            if me["role"] == "admin":
                cur.execute("SELECT id FROM stations WHERE brand=?", (get_brand(),))
                return {int(r["id"]) for r in cur.fetchall()}
            if me["role"] == "jefe_estacion":
                sid0 = me.get("station_id")
                cur.execute("SELECT group_name FROM stations WHERE id=? AND brand=?", (sid0, get_brand()))
                r = cur.fetchone()
                g = (r["group_name"] if r else None)
                if g:
                    cur.execute("SELECT id FROM stations WHERE brand=? AND group_name=?", (get_brand(), g,))
                    return {int(x["id"]) for x in cur.fetchall()}
                return {int(sid0)} if sid0 else set()
            # operador/contador/auditor: solo su estación por defecto
            sid0 = me.get("station_id")
            return {int(sid0)} if sid0 else set()

        allowed = allowed_station_ids()
        if not station_id:
            station_id = next(iter(allowed), None)

        if not station_id or int(station_id) not in allowed:
            conn.close()
            return jsonify({"error": "forbidden_station"}), 403

        sid = int(station_id)

        today = datetime.date.today()
        today_str = today.isoformat()
        month_start = today.replace(day=1).isoformat()
        year_start = today.replace(month=1, day=1).isoformat()

        def calc(range_start: str, range_end: str | None):
            if range_end:
                cur.execute(
                    "SELECT COUNT(*) AS c FROM calendar_events WHERE brand=? AND (station_id IS NULL OR station_id=?) AND start_date>=? AND start_date<=?",
                    (get_brand(), sid, range_start, range_end),
                )
            else:
                cur.execute(
                    "SELECT COUNT(*) AS c FROM calendar_events WHERE brand=? AND (station_id IS NULL OR station_id=?) AND start_date>=?",
                    (get_brand(), sid, range_start),
                )
            total = int(cur.fetchone()["c"])

            if range_end:
                cur.execute(
                    "SELECT COUNT(DISTINCT event_id) AS c FROM submissions WHERE brand=? AND station_id=? AND status IN ('submitted','approved') "
                    "AND event_id IN (SELECT id FROM calendar_events WHERE brand=? AND (station_id IS NULL OR station_id=?) AND start_date>=? AND start_date<=?)",
                    (get_brand(), sid, get_brand(), sid, range_start, range_end),
                )
            else:
                cur.execute(
                    "SELECT COUNT(DISTINCT event_id) AS c FROM submissions WHERE brand=? AND station_id=? AND status IN ('submitted','approved') "
                    "AND event_id IN (SELECT id FROM calendar_events WHERE brand=? AND (station_id IS NULL OR station_id=?) AND start_date>=?)",
                    (get_brand(), sid, get_brand(), sid, range_start),
                )
            done = int(cur.fetchone()["c"])
            pct = int(round((done / total) * 100)) if total else 0
            return {"done": done, "total": total, "pct": pct}

        daily = calc(today_str, today_str)
        monthly = calc(month_start, today_str)
        yearly = calc(year_start, today_str)

        # Overdue = eventos pasados sin evidencia enviada/aprobada
        cur.execute(
            "SELECT COUNT(*) AS c FROM calendar_events WHERE brand=? AND (station_id IS NULL OR station_id=?) AND start_date < ?",
            (get_brand(), sid, today_str),
        )
        total_past = int(cur.fetchone()["c"])
        cur.execute(
            "SELECT COUNT(DISTINCT event_id) AS c FROM submissions WHERE brand=? AND station_id=? AND status IN ('submitted','approved') "
            "AND event_id IN (SELECT id FROM calendar_events WHERE brand=? AND (station_id IS NULL OR station_id=?) AND start_date < ?)",
            (get_brand(), sid, get_brand(), sid, today_str),
        )
        done_past = int(cur.fetchone()["c"])
        overdue = max(total_past - done_past, 0)

        conn.close()
        return jsonify({"station_id": sid, "today": today_str, "daily": daily, "monthly": monthly, "yearly": yearly, "overdue": overdue})
    # Blueprint registration
    app.register_blueprint(activities_bp)
